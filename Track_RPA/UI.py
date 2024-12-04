#Project Header

# Python Libraries
# Selenium
from selenium import webdriver #Web Automation Library
from selenium.webdriver.common.by import By #XPATH Web Selection Library
from selenium.webdriver.chrome.service import Service #Web Automation Library
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.ui import Select #ComboBox Web Selection Library
from selenium.webdriver.chrome.options import Options #Change Language Browser
# Email
import smtplib #Library for Emails
from email.mime.text import MIMEText #Library for Emails
from email.mime.multipart import MIMEMultipart #Library for Emails
# Graphic Interface
from tkinter import * #Graphical Interface Library
from tkinter import Tk, ttk, messagebox #Graphical Interface Library
from tkinter.filedialog import askopenfilename #Library to Select Machine files
from ttkthemes import ThemedStyle
import customtkinter
import awesometkinter as atk
from PIL import Image, ImageTk #Library for Images with TkInter
# Time
import time
from time import sleep #Python Timeout Library
from datetime import datetime
# Utilities
from winotify import Notification, audio # Notifications Windows Library
import random as r # Randomize Number Library
import pandas as pd # Library for Working with files
import openpyxl # Excel Integration Library
import psutil
import win32com.client as win32
import os # Windows Path Library
import requests # Library for API
import sys # System Library
# Write PDF
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import A4
# Data Base
import sqlite3
import psycopg2

class Functions():
    # Create Data Base
    def TWBD(self):
        self.client_uuid = "78ee5ef4fd5eeb96eadd00a690581c4a"

        # Establish the connection to the database
        self.conn = psycopg2.connect(
            dbname="Track_RPA",
            user="postgres",
            password="senha21890",
            host="localhost",
            port="5432"
        )

        # Create a cursor to run queries
        self.sql = self.conn.cursor()

    # Call Screen Functions
    # Button Go to Back
    def call_back(self):
        # Deletes all Elements from Frame 01 and recalls other Elements from Screen 02
        for self.Widgets in self.FR1.winfo_children():  # Selects the Child Elements of Frame_01, that is, Labels, Inputs and Buttons.
            self.Widgets.destroy()  # Destroy all Selected.

        # Checks the page the user is on to return to the previous one.
        if self.LO == "Screen_File":
            self.LO = None
            self.file = None
            self.Screen_Login()
        elif self.LO == "Screen_Exe":
            self.LO = None
            self.file = None
            self.Screen_File()
        elif self.LO == "Screen_Help":
            self.LO = None
            self.Screen_Login()

    # Button for Help User
    def call_help(self):

        # Deletes all Elements from Frame 01 and recalls other Elements from Screen 02
        for self.Widgets in self.FR1.winfo_children():  # Selects the Child Elements of Frame_01, that is, Labels, Inputs and Buttons.
            self.Widgets.destroy()  # Destroy all Selected.
        self.Screen_Help()  # Calls the Help Screen Function.

    # Check Login
    def check_login(self):
        # Receives the Value of the User Field
        if self.en_user.get():
            self.Login = self.en_user.get()

        # Receive the Value of the Pass Field
        if self.en_pass.get():
            self.Pass = self.en_pass.get()

        # Checks whether the User is registered in the Users Table..
        self.sql.execute("SELECT * FROM tb_users WHERE email='" + self.Login + "'")
        record = self.sql.fetchone()
        if record:
            Email = record[2]
            if self.Login == Email:
                # Checks whether the Portal in which the user is registered has been registered in the WebPortal Table.
                self.Link = record[1]
                self.sql.execute("SELECT * FROM tb_webportal WHERE Link='" + self.Link + "'")
                record = self.sql.fetchone()
                if record:
                    self.API = record[3]
                    self.UUID = record[4]
                    self.Environment = record[5]
                    if self.API == "":
                        messagebox.showwarning("Login", "This API Token is not Registered!")
                        return
                    else:
                        self.sql.execute("SELECT * FROM tb_webportal WHERE API_Token='" + self.API + "'")
                        record = self.sql.fetchone()
                        if record:

                            if self.Environment == "Test":
                                POST_URL = "https://api.test.tracktraceweb.com/2.0/user/login"  # TT2 API Path
                            elif self.Environment == "Production":
                                POST_URL = "https://api.tracktraceweb.com/2.0/user/login" # TT2 API Path

                            self.headers = {'Content-type': 'application/x-www-form-urlencoded',
                                            'Accept': 'application/json',
                                            'Authorization': self.API}

                            inf = {  # Variable that stores the required API fields in a JSON model
                                "client_uuid": self.UUID,
                                "username": self.Login,
                                "password": self.Pass
                            }

                            response = requests.post(POST_URL, data=inf, headers=self.headers)  # Assemble the Request and Send by Post

                            if response.status_code == 200:  # If the result is Request OK, that is, 200

                                # Deletes all Elements from Frame 01 and recalls other Elements from Screen 02
                                for self.Widgets in self.FR1.winfo_children():  # Selects the Child Elements of Frame_01, that is, Labels, Inputs and Buttons.
                                    self.Widgets.destroy()  # Destroy all Selected.
                                self.Screen_File()  # Calls the Next Screen Function.

                            else:  ##If the result is Request NOK, that is, 200
                                print(response)
                                messagebox.showwarning("Login", "There was an error logging into your portal. Please check with your Track RPA administrator!")
                                return
                else:
                    messagebox.showwarning("Login", "The Link linked to this user is not registered in Our RPA. Please, contact support!")
                    return
        else:
            messagebox.showwarning("Login", "This User is not linked to any portal!")
            return

    # Receive CSV file
    def get_file(self):
        # Checks if the Excel File that Feeds the RPA is Open
        self.file = askopenfilename(filetypes=(("CSV files", "*.xlsx"),))

        def normalize_path(file_path):
            return os.path.normcase(os.path.normpath(file_path))

        self.normalized_file_path = normalize_path(self.file)

        for process in psutil.process_iter():
            try:
                if 'excel' in process.name().lower():
                    # Obtém os arquivos abertos pelo processo
                    open_files = process.open_files()
                    for open_file in open_files:
                        if self.normalized_file_path == open_file.path.lower():
                            self.file = None
                            self.normalized_file_path = None
                            messagebox.showwarning("Impossible to Process!", "Please close the RPA Processing File to continue.")
                            return
            except psutil.Error:
                print("Error")

        if self.file:
            name = os.path.basename(self.file)
            Tb = pd.read_excel(self.file)
            self.cont = len(Tb)
            Tb = None

            if name != "Test_Big.xlsx":
                messagebox.showwarning("TTRX - RPA", "Incorrect File Selected!")
                return

        else:
            return

        for self.Widgets in self.FR1.winfo_children():  # Selects the Child Elements of Frame_01, that is, Labels, Inputs and Buttons.
            self.Widgets.destroy()  # Destroy all Selected.
        self.generate_lots()
        self.Screen_Execute()

    # Calls each function in the Code
    def call_functions(self):
        def create_file(body):
            ph_no = []
            # A loop is used to generate 5 Numbers.
            for i in range(0, 5):
                ph_no.append(r.randint(0, 9))  # Each Number from 0 to 9
            # I take the numbers from the array and put them together
            Log = ''
            for i in ph_no:
                Log += str(i)

            # Specify the path to the desired folder
            Path = getattr(sys, '_MEIPASS', os.path.abspath(os.path.dirname(__file__)))
            Exe = Path + "\\Archives\\Log"  # Stores the path where the program is running

            # Check if the folder exists, if not, create it
            if not os.path.exists(Exe):
                os.makedirs(Path)

            # Match the folder path to the error file name
            caminho_arquivo_erro = os.path.join(Exe, 'Erro_' + Log + '.txt')

            # Open a file to write errors
            with open(caminho_arquivo_erro, 'w') as f:
                f.writelines(body)

        def send_email(outlook_email, outlook_pass, recipient, subject, body):
            # SMTP Outlook Configure
            servidor_smtp = "smtp.office365.com"
            porta_smtp = 587

            # Create Email
            mensagem = MIMEMultipart()
            mensagem['From'] = outlook_email
            mensagem['To'] = recipient
            mensagem['Subject'] = subject

            # Add Body in Message
            mensagem.attach(MIMEText(body, 'plain'))

            # Configure the SMTP connection
            try:
                servidor = smtplib.SMTP(servidor_smtp, porta_smtp)
                servidor.starttls()
                servidor.login(outlook_email, outlook_pass)

                # Sending Email
                servidor.sendmail(outlook_email, recipient, mensagem.as_string())

                print("Email Successfully Sent!")

            except Exception as e:
                print("Error in Sending Email:", str(e))

        try:

            #self.Trading_Proc = "Trading_98860"

            self.header()

            def adicionar_informacoes_arquivo(file_name, informacoes):
                # Especifique o caminho para a pasta desejada
                Path = getattr(sys, '_MEIPASS', os.path.abspath(os.path.dirname(__file__)))
                Exe = os.path.join(self.Path + "\\Archives\\Output")

                # Verifique se a pasta existe, se não, crie-a
                if not os.path.exists(Exe):
                    os.makedirs(Path)

                # Combine o caminho da pasta com o nome do arquivo de erro
                path_file_output = os.path.join(Exe, file_name)

                try:
                    with open(path_file_output, 'a') as arquivo:
                        arquivo.write(informacoes + '\n')
                    print(f"Information successfully added to the file {file_name}.")
                except Exception as e:
                    print(f"Error adding information to file: {str(e)}")

            # Generates a random number of Output File
            ph_no = []
            # A loop is used to generate 5 Numbers.
            for i in range(0, 5):
                ph_no.append(r.randint(0, 9))  # Each Number from 0 to 9
            # I take the numbers from the array and put them together
            Arq = ''
            for i in ph_no:
                Arq += str(i)

            file_name = 'Output_' + Arq + '.txt'
            self.f = file_name

            Start_Time_Process = datetime.now().strftime("%I:%M %p")
            Date_Process = datetime.now().strftime("%m/%d/%Y")

            informacoes_para_adicionar = "The Test Process executed by TrackTraceRX started at " + Start_Time_Process + " on \n" + Date_Process + ".\n"
            adicionar_informacoes_arquivo(file_name, informacoes_para_adicionar)

            # FILE THAT READS THE FUNCTIONS THAT MUST BE CALLED
            Path = getattr(sys, '_MEIPASS', os.path.abspath(os.path.dirname(__file__)))
            cam = os.path.join(Path + '\\Archives', 'config.txt')

            try:
                # Retrieves Table Values
                self.sql.execute("SELECT * FROM tb_module WHERE Link='" + self.Link + "'")
                record = self.sql.fetchone()

                if record:
                    vr = record[2]
                    if vr == "T":
                        self.tranding_partners()  # Partner Creation Function

                        informacoes_para_adicionar = "We started the " + self.Module + " Module Test Process at " + self.Start_Time_Module + ", and concluded " + self.Result + " \nat " + self.End_Time_Module + ". Trading Partner " + self.Trading_Proc + " was Created.\n"
                        adicionar_informacoes_arquivo(file_name, informacoes_para_adicionar)
                    vr = record[3]
                    if vr == "T":
                        self.product_management()  # Product Creation Function

                        informacoes_para_adicionar = "\n We started the " + self.Module + " Module Test Process at " + self.Start_Time_Module + ", and concluded " + self.Result + " \nat " + self.End_Time_Module + "."
                        adicionar_informacoes_arquivo(file_name, informacoes_para_adicionar)
                    vr = record[4]
                    if vr == "T":
                        self.receiving_serials()  # Receiving Product Function

                        informacoes_para_adicionar = "\n We started the " + self.Module + " Module Test Process at " + self.Start_Time_Module + ", and concluded " + self.Result + " at \n" + self.End_Time_Module + ". Purchase Order " + self.PO_Name_Serial + " was Created."
                        adicionar_informacoes_arquivo(file_name, informacoes_para_adicionar)
                    vr = record[5]
                    if vr == "T":
                        self.outbound()  # Sales Order Function

                        informacoes_para_adicionar = "\n We started the " + self.Module + " Module Test Process at " + self.Start_Time_Module + ", and concluded " + self.Result + " at \n" + self.End_Time_Module + ". Sales Order " + self.SO_Name_Serial + " was Created."
                        adicionar_informacoes_arquivo(file_name, informacoes_para_adicionar)
                    vr = record[6]
                    if vr == "T":
                        self.outbound_by_picking()  # Sales Order By Picking Function

                        informacoes_para_adicionar = "\n We started the " + self.Module + " Module Test Process at " + self.Start_Time_Module + ", and concluded \n" + self.Result + " at " + self.End_Time_Module + ". Sales Order by Picking" + self.SO_Name_SBP + " was Created."
                        adicionar_informacoes_arquivo(file_name, informacoes_para_adicionar)

                        self.RMA() #RMA Function

                        informacoes_para_adicionar = "\n We started the " + self.Module + " Module Test Process at " + self.Start_Time_Module + ", and concluded " + self.Result + " at \n" + self.End_Time_Module + ". RMA for Sales Order by Picking" + self.SO_Name_SBP + " was Created."
                        adicionar_informacoes_arquivo(file_name, informacoes_para_adicionar)
                    vr = record[7]
                    if vr == "T":
                        self.containers()  # Containers Creation Function

                        informacoes_para_adicionar = "\n We started the " + self.Module + " Module Test Process at " + self.Start_Time_Module + ", and concluded " + self.Result + "\nat " + self.End_Time_Module + ". Container urn:epc:id:sscc:0245869.11111111" + self.Container_URN + " was Created."
                        adicionar_informacoes_arquivo(file_name, informacoes_para_adicionar)
                    vr = record[8]
                    if vr == "T":
                        self.quarantine()  # Quarantine Function

                        informacoes_para_adicionar = "\n We started the " + self.Module + " Module Test Process at " + self.Start_Time_Module + ", and concluded " + self.Result + "\nat " + self.End_Time_Module + "."
                        adicionar_informacoes_arquivo(file_name, informacoes_para_adicionar)
                    vr = record[10]
                    if vr == "T":
                        self.com_pack()  # Comission, Packaging and Unpacking Function

                        informacoes_para_adicionar = "\n We started the " + self.Module + " Module Test Process at " + self.Start_Time_Module + ", and \nconcluded " + self.Result + " at " + self.End_Time_Module + "."
                        adicionar_informacoes_arquivo(file_name, informacoes_para_adicionar)
                    vr = record[11]
                    if vr == "T":
                        self.disporsal()  # Inventory Function

                        informacoes_para_adicionar = "\n We started the " + self.Module + " Module Test Process at " + self.Start_Time_Module + ", and concluded " + self.Result + "\nat " + self.End_Time_Module + "."
                        adicionar_informacoes_arquivo(file_name, informacoes_para_adicionar)
                    vr = record[12]
                    if vr == "T":
                        self.transformation()  # Transformation Function

                        informacoes_para_adicionar = "\n We started the " + self.Module + " Module Test Process at " + self.Start_Time_Module + ", and concluded " + self.Result + "\nat " + self.End_Time_Module + "."
                        adicionar_informacoes_arquivo(file_name, informacoes_para_adicionar)
                    vr = record[13]
                    if vr == "T":
                        self.company_management()  # Company Function

                        informacoes_para_adicionar = "\n We started the " + self.Module + " Module Test Process at " + self.Start_Time_Module + ", and concluded " + self.Result + " at " + self.End_Time_Module + "."
                        adicionar_informacoes_arquivo(file_name, informacoes_para_adicionar)

            except Exception as e:
                print(f"Error executing methods: {str(e)}")

                # Send email if an exception occurs
                outlook_email = "vbacacicci@outlook.com"
                outlook_pass = "senha21890"
                recipient = "victor.bacacicci@outlook.com"
                subject = "Error Found in TrackRPA"
                body = f"The following error occurred: {str(e)}"

                create_file(body)
                send_email(outlook_email, outlook_pass, recipient, subject, body)

                messagebox.showwarning("TTRX - RPA", "We had a problem processing your file, but don't worry, an email has already been sent to the responsible team to analyze your error and they will contact you shortly!")

                # Deletes all Elements from Frame 01 and recalls other Elements from Screen 02
                for self.Widgets in self.FR1.winfo_children():  # Selects the Child Elements of Frame_01, that is, Labels, Inputs and Buttons.
                    self.Widgets.destroy()  # Destroy all Selected.
                self.file = None
                self.LO = "Screen_File"
                self.Screen_File()

        except Exception as e:

            # Send email if an exception occurs
            outlook_email = "vbacacicci@outlook.com"
            outlook_pass = "senha21890"
            recipient = "victor.bacacicci@outlook.com"
            subject = "Error Found in TrackRPA"
            body = f"The following error occurred: {str(e)}"

            create_file(body)
            send_email(outlook_email, outlook_pass, recipient, subject, body)

            messagebox.showwarning("TTRX - RPA", "We had a problem processing your file, but don't worry, an email has already been sent to the responsible team to analyze your error and they will contact you shortly!")

            # Deletes all Elements from Frame 01 and recalls other Elements from Screen 02
            for self.Widgets in self.FR1.winfo_children():  # Selects the Child Elements of Frame_01, that is, Labels, Inputs and Buttons.
                self.Widgets.destroy()  # Destroy all Selected.
            self.LO = "Screen_02"
            self.Screen_File()

        self.pdf()
        self.navegador.quit()

        # Deletes all Elements from Frame 01 and recalls other Elements from Screen 02
        for self.Widgets in self.FR1.winfo_children():  # Selects the Child Elements of Frame_01, that is, Labels, Inputs and Buttons.
            self.Widgets.destroy()  # Destroy all Selected.
        self.file = None
        self.normalized_file_path = None
        self.LO = "Screen_File"
        self.Screen_File()  # Calls the Help Screen Function.
        messagebox.showwarning("TTRX - RPA", "Your File Has Been Processed Successfully! If you want, you can process it again when performing a new upload!")

    # Generates Transaction Serials
    def generate_serials(self):
        # GENERATING SERIAL file
        # Function to generate random numbers with 9 digits
        def gerar_numero():
            return r.randint(100000000, 999999999)
            # Create a new Excel file

        wb = openpyxl.Workbook()
        sheet = wb.active

        # Add Column Names
        sheet.append(["item", "serial"])


        Q = int(self.QUANT) + 1
        # Generate seven random numbers with IDs and store them in Excel file
        for i in range(1, Q):
            numero_aleatorio = gerar_numero()  # Call the Function that Generates the Number
            sheet.append([i, numero_aleatorio])  # Save the Item Number (1 to 7) in the First Column and the created serial in the second column

        Path = getattr(sys, '_MEIPASS', os.path.abspath(os.path.dirname(__file__)))
        Exe = Path + "\\Archives\\Serials.xlsx"  # Stores the path where the program is running

        # Save the Excel file
        wb.save(Exe)

        #Automation function

    # Generates Transaction Lots
    def generate_lots(self):
        Tb = pd.read_excel(self.file)
        for i, product in enumerate(Tb["product_name"]):
            item = Tb.loc[i, "item"]

            work = openpyxl.load_workbook(self.file)
            sheet = work["Planilha1"]

            linha = item + 1
            coluna = 6

            ph_no = []
            # A loop is used to generate 5 Numbers.
            for i in range(0, 5):
                ph_no.append(r.randint(0, 9))  # Each Number from 0 to 9
                # I take the numbers from the array and put them together
            L = ''
            for i in ph_no:
                L += str(i)

            sheet.cell(row=linha, column=coluna, value="LOTSERIAL" + L)
            work.save(self.file)

    # Selenium Connection and Login Portal
    def header(self):

        # SELENIUM VARIABLES
        chrome_service = Service("C:\\Users\\Victor Angêlo\\OneDrive\\TRACK\\Development\\Python Development\\Automation\\Track_RPA\\Archives\\chromedriver.exe")  # ChromeDriver Location
        self.navegador_options = Options()
        self.navegador_options.add_argument('--lang=en-US') #Change Language Browser
        self.navegador = webdriver.Chrome(service=chrome_service, options=self.navegador_options)  # Open Browser
        self.navegador.maximize_window()  # Open Browser in Full Screen

        #Open Dashboard.
        print(self.Link)
        self.navegador.get(self.Link) #Link

        #Check if you need to login.
        while len(self.navegador.find_elements(By.ID, 'login_form_img_preload')) == 1:

            Email = self.Login
            Senha = self.Pass

            # Start of Authentication

            try:
                botao = WebDriverWait(self.navegador, 10).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="auth__login_form__username"]'))
                )
                self.navegador.find_element(By.XPATH, '//*[@id="auth__login_form__username"]').send_keys(Email)  # Email Field
                sleep(2)
            except:
                print("A")

            try:
                botao = WebDriverWait(self.navegador, 10).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="auth__login_form__step1_next_btn"]'))
                )
                self.navegador.find_element(By.XPATH, '//*[@id="auth__login_form__step1_next_btn"]').click()  # Next Button
                sleep(2)
            except:
                print("A")

            try:
                botao = WebDriverWait(self.navegador, 10).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="auth__login_form__password"]'))
                )
                self.navegador.find_element(By.XPATH, '//*[@id="auth__login_form__password"]').send_keys(Senha)  # Pass Field
                sleep(2)
            except:
                print("A")

            try:
                botao = WebDriverWait(self.navegador, 10).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="auth__login_form__step2_next_btn"]'))
                )
                self.navegador.find_element(By.XPATH, '//*[@id="auth__login_form__step2_next_btn"]').click()  # Login Button
                sleep(5)
            except:
                print("A")

    # Functions of Screen Builds
    def Screen_Login(self):
        self.LO = "Screen_Login"

        # Frames in Screen 01
        estilo = ttk.Style()
        estilo.configure("TFrame", bd=0, background="#f8f8f8", relief="groove", padding=(10, 10))
        self.FR1 = ttk.Frame(self.Screen, style="TFrame")
        self.FR1.place(relx=0.3, rely=0.35, relwidth=0.4, relheight=0.4)

        # Label User
        self.lb_user = Label(self.FR1, background='#f8f8f8', foreground='#729898', text="Username", font=('Sans-serif', 8, 'bold')).place(relx=0.38, rely=0.15)

        # Label Pass
        self.lb_pass = Label(self.FR1, background='#f8f8f8', foreground='#729898', text="Password", font=('Sans-serif', 8, 'bold')).place(relx=0.38, rely=0.40)

        # Input User
        self.en_user = ttk.Entry(self.FR1)
        self.en_user.place(relx=0.15, rely=0.25, relwidth=0.70)

        # Input Pass
        self.en_pass = ttk.Entry(self.FR1, show="*")
        self.en_pass.place(relx=0.15, rely=0.50, relwidth=0.70)

        # Button Style
        estilo = ttk.Style()
        estilo.configure("bt_login.TButton", foreground='#729898', font=('Sans-serif', 8))
        # Button Login
        self.bt_login = ttk.Button(self.FR1, style="bt_login.TButton", text="Next", command=self.check_login)
        self.bt_login.pack(pady=20, padx=50)
        self.bt_login.place(relx=0.27, rely=0.68, relwidth=0.45)

        # Link Help
        link_help = ttk.Label(self.FR1, text="Help", cursor="hand2", background='#f8f8f8', foreground='#729898', font=('Sans-serif', 8))
        link_help.place(relx=0.45, rely=0.85)
        # Adicionar um evento de clique ao rótulo
        link_help.bind("<Button-1>", lambda event: self.call_help())

        # Baseboard
        self.lb_base = Label(self.FR2, background='#072144', text="Powered By TrackTraceRX", fg='white', font=('Sans-serif', 8, 'bold')).place(relx=0.38, rely=0.01)

    def Screen_File(self):
        self.LO = "Screen_File"

        self.lb_main = Label(self.FR1, background='#f8f8f8', bd=4, fg='#729898', font=('Sans-serif', 8, 'bold'), text="Welcome to TTRX - RPA").place(relx=0.24, rely=0.15)
        self.lb_choose = Label(self.FR1, background='#f8f8f8', bd=4, fg='#729898', font=('Sans-serif', 8, 'bold'), text="Choose your file for Processing").place(relx=0.18, rely=0.40)

        # Button Style
        estilo = ttk.Style()
        estilo.configure("TButton", foreground='#729898', font=('Sans-serif', 8))
        # Button Login
        self.bt_choose = ttk.Button(self.FR1, style="bt_login.TButton", text="Browse file", command=self.get_file)
        self.bt_choose.place(relx=0.28, rely=0.65, relwidth=0.45)

        # Link Back
        link_help = ttk.Label(self.FR1, text="LogOut", cursor="hand2", background='#f8f8f8', foreground='#729898', font=('Sans-serif', 8))
        link_help.place(relx=0.02, rely=0.87)
        # Add a click event to the label
        link_help.bind("<Button-1>", lambda event: self.call_back())

        # Link Back
        link_help = ttk.Label(self.FR1, text="- Help", cursor="hand2", background='#f8f8f8', foreground='#729898', font=('Sans-serif', 8))
        link_help.place(relx=0.16, rely=0.87)
        # Add a click event to the label
        link_help.bind("<Button-1>", lambda event: self.call_back())

        # Baseboard
        self.lb_base = Label(self.FR2, background='#072144', text="Powered By TrackTraceRX", fg='white', font=('Sans-serif', 8, 'bold')).place(relx=0.38, rely=0.01)

    def Screen_Execute(self):
        self.LO = "Screen_Exe"

        # Deletes all Elements from Frame 01 and recalls other Elements from Screen 02
        for self.Widgets in self.FR1.winfo_children():  # Selects the Child Elements of Frame_01, that is, Labels, Inputs and Buttons.
            self.Widgets.destroy()  # Destroy all Selected.

        self.lb_main1 = Label(self.FR1, text="Welcome to TTRX - RPA", background='#f8f8f8', bd=4, fg='#729898', font=('Sans-serif', 8, 'bold')).place(relx=0.24, rely=0.15)
        self.lb_main2 = Label(self.FR1, text="Wait! We are Processing your File.", background='#f8f8f8', bd=4, fg='#729898', font=('Sans-serif', 8, 'bold')).place(relx=0.15, rely=0.50)

        self.Screen.update()  # Updates the graphical interface to start the process
        sleep(2)
        self.call_functions()

    def Screen_Help(self):
        self.LO = "Screen_Help"

        self.lb_instruction1 = Label(self.FR1, background='#f8f8f8', bd=4, fg='#729898', font=('Sans-serif', 8, 'bold'), text="This product was designed to function \n  as a process facilitator within the \n TrackTrace2 Portal. To use it, you will \n need to have an RPA User registered \n in the Staff Management of your Portal.").place(relx=0.1, rely=0.10)
        self.lb_instruction2 = Label(self.FR1, background='#f8f8f8', bd=4, fg='#729898', font=('Sans-serif', 8, 'bold'), text="Our support team can help with any \n other questions you may have.").place(relx=0.13, rely=0.60)

        image_path = os.path.join(self.Path + "\\Archives\\images\\", "back_icon.png")
        imagem = Image.open(image_path)
        imagem = imagem.resize((13, 13))
        imagem_tk = ImageTk.PhotoImage(imagem)

        self.bt_back = Button(self.FR1, background='#f8f8f8', bd=0, image = imagem_tk, command=self.call_back, compound=CENTER)
        self.bt_back.image = imagem_tk
        self.bt_back.place(relx=0.01, rely=0.85, relwidth=0.13, relheight=0.13)

        # Baseboard
        self.lb_base = Label(self.FR2, background='#072144', text="Powered By TrackTraceRX", fg='white', font=('Sans-serif', 8, 'bold')).place(relx=0.38, rely=0.01)

    # Functions for each Portal Module
    def tranding_partners(self):

        self.Module = "Trading Partners"
        self.Start_Time_Module = datetime.now().strftime("%I:%M %p")

        # START OF THE "TRANDING PARTNERS" SCREEN

        #Open Trading Partners Screen.
        self.navegador.get(self.Link + "/trading_partners/")  # Trading Partners Link
        sleep(2)

        #Add TP
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[2]/div/div[3]/span').click() #Button Add
        sleep(2)

        #General Tab
        ph_no = []
        #A loop is used to generate 5 Numbers.
        for i in range(0, 5):
            ph_no.append(r.randint(0, 9))  #Each Number from 0 to 9
        #I take the numbers from the array and put them together
        TP = ''
        for i in ph_no:
            TP += str(i)
        self.Trading_Proc = "Trading_" + TP
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[3]/div[2]/input').send_keys(self.Trading_Proc) #TP Name Field
        sleep(2)

        #Generating the GLN Company Prefix
        ph_no = []
        #A loop is used to generate 07 Numbers.
        for i in range(0, 7):
            ph_no.append(r.randint(0, 9))  #The other numbers are between 0 and 9.
        #I take the numbers from the array and put them together
        CPGLN = ''
        for i in ph_no:
            CPGLN += str(i)

        #Generating the GLN Sub Main
        ph_no = []
        #A loop is used to generate 07 Numbers.
        for i in range(0, 5):
            ph_no.append(r.randint(0, 9))  #The other numbers are between 0 and 9.
        #I take the numbers from the array and put them together
        SUBGLN = ''
        for i in ph_no:
            SUBGLN += str(i)

        #Generating the GLN Verifier Digit
        VLGLN = r.randrange(1, 9)

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[7]/div[2]/input').send_keys(CPGLN + SUBGLN) #GLN Field
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[8]/div[2]/input').send_keys(CPGLN)  #Company Prefix Field
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[9]/div[2]/input').send_keys("urn:epc:id:sgln:" + CPGLN + "." + SUBGLN + "." + str(VLGLN))  #SGLN Field

        #Contacts Tab
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[2]/span').click()  #Contacts Tab Click
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[1]/div[2]/div[2]/input').send_keys("teste@teste.com")  #Email Notification Field

        #Address Tab
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[3]/span').click()  #Address Tab
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click()  #Save TP
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[3]/div/div[2]/div/div[1]/span').click()  #Button Add Address
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[3]/div[2]/input').send_keys(self.Trading_Proc + " Main")  #Name of Address
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[4]/div[2]/input').send_keys(CPGLN + SUBGLN + str(VLGLN))  #GLN
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[5]/div[2]/input').send_keys("urn:epc:id:sgln:" + CPGLN + "." + SUBGLN + "." + str(VLGLN))  #SGLN
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[6]/div[2]/input').send_keys(self.Trading_Proc)  #Recipient Name
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[7]/div[2]/input').send_keys("Teste")  #Line Address 01
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[13]/div[2]/input').send_keys("Teste")  #City
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[14]/div[2]/input').send_keys("Teste")  #ZIP Code
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click() #TP Address Add Button
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click() #Main Form Add Button

        self.End_Time_Module = datetime.now().strftime("%I:%M %p")
        self.Result = "Successfully"
        print(self.Trading_Proc + " Complete!")

    def product_management(self):

        self.Module = "Product Management"
        self.Start_Time_Module = datetime.now().strftime("%I:%M %p")

        #START OF THE "MANAGEMENT PRODUCTS" SCREEN

        # Open the Table with Products
        Tb = pd.read_excel(self.file)
        for l, product in enumerate(Tb["product_name"]):
            NDCOFICIAL = Tb.loc[l, "ndc"]
            GS1CP = Tb.loc[l, "gs1_item_company_prefix"]
            GS1ITEM = Tb.loc[l, "gs1_item_reference"]
            NAME = Tb.loc[l, "product_name"]

            if self.Environment == "Test":
                GET_URL = 'https://api.test.tracktraceweb.com/2.0/products?identifier_us_ndc=' + NDCOFICIAL  # MAKE A GET WITH NDC
            elif self.Environment == "Production":
                GET_URL = 'https://api.tracktraceweb.com/2.0/products?identifier_us_ndc=' + NDCOFICIAL  # MAKE A GET WITH NDC

            # LINE THAT EXECUTES THE API WITH THE REQUESTS LIBRARY
            response = requests.get(GET_URL, headers=self.headers)

            if response.status_code == 200:  # If the API Response is Positive
                response_dic = response.json()  # Create a JSON Dictionary with the POST Response
                RA = response_dic['nb_total_results']  # Stores the requested JSON tag in the variable!
                if RA < 1:
                    # Create Product if not Exist
                    if self.Environment == "Test":
                        POST_URL = "https://api.test.tracktraceweb.com/2.0/products"  # TT2 API Path
                    elif self.Environment == "Production":
                        POST_URL = "https://api.tracktraceweb.com/2.0/products"  # TT2 API Path

                    inf = {  # Variable that stores the required API fields in a JSON model
                        "type": "Pharmaceutical",
                        "status": "AVAILABLE",
                        "gs1_company_prefix": GS1CP,
                        "gs1_id": GS1ITEM,
                        "is_active": 'true',
                        "product_descriptions": "[{\"language_code\": \"en\",\"name\": \"" + NAME + "\",\"description\": \"" + NAME + "\",\"product_long_name\": \"" + NAME + "\"}]"
                    }

                    response = requests.post(POST_URL, data=inf, headers=self.headers)  # Assemble the Request and Send by Post
                    response_dic = response.json()  # Create a JSON Dictionary with the POST Response
                    UUID = response_dic['uuid']  # Stores the requested JSON tag in the variable!
                    print(UUID)

                    # If the product has been added
                    if response.status_code == 200:  # If the API Response is Positive
                        # Add NDC in Product Created
                        if self.Environment == "Test":
                            POST_URL = "https://api.test.tracktraceweb.com/2.0/products/" + UUID + "/identifiers"  # TT2 API Path
                        elif self.Environment == "Production":
                            POST_URL = "https://api.tracktraceweb.com/2.0/products/" + UUID + "/identifiers"  # TT2 API Path

                        inf = {  # Variable that stores the required API fields in a JSON model
                            "product_uuid": UUID,
                            "identifier_code": "US_NDC",
                            "value": NDCOFICIAL
                        }

                        response = requests.post(POST_URL, data=inf, headers=self.headers)  # Assemble the Request and Send by Post
                        response_dic = response.json()  # Create a JSON Dictionary with the POST Response
                        print(response_dic)
            else:
                print("Error in API")

        self.End_Time_Module = datetime.now().strftime("%I:%M %p")
        self.Result = "Successfully"

    def receiving_serials(self):

        self.Module = "Receiving"
        self.Start_Time_Module = datetime.now().strftime("%I:%M %p")

        #START OF THE "RECEIVING/MANUAL INBOUND" SCREEN

        #Receiving
        self.navegador.get(self.Link + "/receiving") #Receiving Link
        sleep(2)

        #Inbound
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[2]/div/div[2]/span').click() #Manual Inbound
        sleep(2)

        #Change Location
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div[1]/div[3]/a').click()  # Change Location
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[1]/form/div[1]/div[2]/div[2]/input').send_keys("Main Location") #Search Location
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[2]/div').click() #Button Search
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[2]/table/tbody/tr/td[3]/div/img').click() #Select Location

        #Change TP
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div[2]/div[3]/a').click() #Change Seller
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[1]/form/div[1]/div[2]/div[2]/input').send_keys(self.Trading_Proc) #Search TP
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[2]/div').click() #Button Search
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[2]/table/tbody/tr[1]/td[3]/div[2]/img').click() #Select TP

        #PO Form
        # Generating PO and SO Identifications
        ph_no = []
        # A loop is used to generate 5 Numbers.
        for i in range(0, 5):
            ph_no.append(r.randint(0, 9))  # Each Number from 0 to 9
            # I take the numbers from the array and put them together
        Purchase_01 = ''
        for i in ph_no:
            Purchase_01 += str(i)
        self.PO_Name_Serial = "PO_Teste_" + Purchase_01

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div[4]/div[2]/input').send_keys(self.PO_Name_Serial) #Customer Order ID
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div[5]/div[2]/input').send_keys(self.PO_Name_Serial) #Invoice
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div[6]/div[2]/input').send_keys(self.PO_Name_Serial) #PO Number
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div[4]/div[4]/input').send_keys(self.PO_Name_Serial) #Internal Reference
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div[5]/div[4]/input').send_keys(self.PO_Name_Serial) #Release
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div[6]/div[4]/input').send_keys(self.PO_Name_Serial) #Order Number

        #Sold By and Ship From Tab
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[2]/span').click()  #Sold By

        #Bought By and Ship To Tab
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[3]/span').click()  #Bought By

        # #Shipment Tab
        # self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[4]/span').click()  # Shipment Tab
        # sleep(2)
        # Select(self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[4]/div[6]/div[2]/div[1]/div[2]/div[3]/div/select')).select_by_index(1)  # Select Lot Select

        #Itens Tab
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[5]').click() #Line Items
        sleep(1)

        Tb = pd.read_excel(self.file)
        for i, product in enumerate(Tb["product_name"]):
            PR = Tb.loc[i, "product_name"]
            NDCOFICIAL = Tb.loc[i, "ndc"]
            EXP_DATE = str(Tb.loc[i, "exp"])
            LOT_NAME = Tb.loc[i, "lot_name"]
            QUANT = str(Tb.loc[i, "quantity_lot"])

            #Enter Serial Based Screen
            self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[5]/div/div[1]/div[2]/div/div/span').click() #Add Product
            sleep(2)
            self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[1]/form/div[1]/div[3]/div[2]/input').send_keys(NDCOFICIAL) #NDC Search
            self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[2]/div[1]').click() #Button to search for informed NDC
            sleep(2)
            self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[2]/table/tbody/tr/td[6]/div[2]/img').click() #Click in Product
            sleep(2)
            self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[3]/div[2]/input').send_keys(QUANT) #Quantity
            self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click() #Ok in Quantity Form
            sleep(2)
            self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[3]/div/div/div[3]/div[2]/div/div/span').click() #Add Lot
            sleep(2)
            self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[1]/div/div/div/div[1]/div[3]/div[2]/input').send_keys(LOT_NAME) #Lot Name
            self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[1]/div/div/div/div[1]/div[4]/div[2]/div[1]/div/input').click() #Radio Serial Based
            self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[1]/div/div/div/div[1]/div[3]/div[4]/input').send_keys(str(EXP_DATE)) #Expiration
            self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[2]/div/button[1]/span').click() #Ok in Add Lot
            sleep(2)
            self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click() #Ok in Product Information
        Tb = None

            # #Enter Lot Based Screen
            # self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[5]/div/div[1]/div[2]/div/div/span').click() #Add Product
            # sleep(2)
            # self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[1]/form/div[1]/div[3]/div[2]/input').send_keys(NDCOFICIAL) #Search NDC
            # self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[2]/div[1]').click() #Button to search for informed NDC
            # sleep(2)
            # self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[2]/table/tbody/tr/td[6]/div[2]/img').click() #Click on the result you brought
            # sleep(2)
            # self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[3]/div[2]/input').send_keys("4") #Enter the Quantity
            # self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click() #Ok in Quantity Form
            # sleep(2)
            # self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[3]/div/div/div[3]/div[2]/div/div/span').click() #Add Lot
            # sleep(2)
            # self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[1]/div/div/div/div[1]/div[3]/div[2]/input').send_keys("TP01PD01LLB01") #Name Lot
            # self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[1]/div/div/div/div[1]/div[4]/div[2]/div[2]/div/input').click() #Radio Lot Based
            # self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[1]/div/div/div/div[1]/div[3]/div[4]/input').send_keys(EXP_DATE) #Expiration
            # self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[2]/div/button[1]/span').click() #Ok in Add Lot
            # sleep(2)
            # self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click() #Ok in Product Information
            # print(PR + " - Lot Based")

        #Aggregation Menu

        #Aggregation Serials
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[6]').click() #Aggregation Tab
        sleep(2)

        Tb = pd.read_excel(self.file) #Adds the Aggregation of all Purchase Order Products
        for l, product in enumerate(Tb["product_name"]):
            item = Tb.loc[l, "item"]
            LOT_NAME = Tb.loc[i, "lot_name"]
            self.QUANT = str(Tb.loc[i, "quantity_lot"])

            self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[6]/div[2]/div/div[2]/div/div/span').click()  # Add Aggregation
            sleep(2)
            self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[4]/div/div/input').click()  # Select Product Aggregation
            Select(self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[5]/div[3]/div/select')).select_by_index(item)  # Select Product Select
            Select(self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[6]/div[3]/div/select')).select_by_index(1)  # Select Lot Select

            self.generate_serials()

            Ab = pd.read_excel(self.Serial)
            for a, serial in enumerate(Ab["serial"]):
                SR = Ab.loc[a, "serial"]
                self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[8]/div[2]/div/div[3]/div[2]/textarea').send_keys(str(SR) + "\n")  # Type inside the Serials TextArea.

            self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click()  # OK in Aggregation
        Tb = None

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click()  #OK in Form Inbound
        sleep(5)

        #Task - Structure that Checks the "Done" Status of the Task
        self.navegador.get(self.Link + "/utilities/background_tasks")
        sleep(2)
        while len(self.navegador.find_elements(By.XPATH, "/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr[1]/td[5]//span[text()='Done']")) != 1:
            self.navegador.refresh()
            sleep(2)

        self.End_Time_Module = datetime.now().strftime("%I:%M %p")
        self.Result = "Successfully"
        print(self.PO_Name_Serial + " Complete!")

    def outbound_by_picking(self):

        self.Module = "Outbound By Picking"
        self.Start_Time_Module = datetime.now().strftime("%I:%M %p")

        #START OF THE "CREATE SALES ORDER/OUTBOUND BY PICKINKG" SCREEN

        self.navegador.get(self.Link)  #Dashboard
        sleep(4)

        #self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[5]/div[2]/div/label').click() #Enter Create Sales Order by Picking
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[6]/div[1]/div/label').click() #Enter Create Sales Order by Picking
        sleep(2)

        #CHANGE TP
        self.navegador.find_element(By.XPATH, '/html/div[2]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[1]/form/div[1]/div[2]/div[2]/input').send_keys(self.Trading_Proc) #Input TP
        self.navegador.find_element(By.XPATH, '/html/div[2]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[2]/div').click() #Search TP
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div[1]/div[2]/div/div/div/div/div/div[2]/table/tbody/tr/td[3]/div/img').click() #Click TP
        sleep(2)

        #CHANGE LOCATION
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[1]/div[3]/a').click() #Search Location
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[1]/form/div[1]/div[2]/div[2]/input').send_keys(self.Location) #Input Search Location
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[2]/div').click() #Click Search
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[2]/table/tbody/tr/td[3]/div/img').click() #Click Location
        sleep(2)

        # Generating PO and SO Identifications
        ph_no = []
        # A loop is used to generate 5 Numbers.
        for i in range(0, 5):
            ph_no.append(r.randint(0, 9))  # Each Number from 0 to 9
        # I take the numbers from the array and put them together
        A = ''
        for i in ph_no:
            A += str(i)
        self.SO_Name_SBP = "SO_Picking_" + A

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[4]/div[2]/input').send_keys(self.SO_Name_SBP) #Customers ID
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[5]/div[2]/input').send_keys(self.SO_Name_SBP) #Invoice
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[6]/div[2]/input').send_keys(self.SO_Name_SBP) #SO number
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[4]/div[4]/input').send_keys(self.SO_Name_SBP) #Internal Reference Nbr
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[5]/div[4]/input').send_keys(self.SO_Name_SBP) #Release Number
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[6]/div[4]/input').send_keys(self.SO_Name_SBP) #Order Number
        Select(self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[7]/div[2]/div/select')).select_by_index(3)  # Select Product

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[5]/span').click() #Change Tab from Pickings
        sleep(2)

        Tb = pd.read_excel(self.file)
        for l, product in enumerate(Tb["product_name"]):
            NDCOFICIAL = Tb.loc[l, "ndc"]
            LOT_NAME = Tb.loc[l, "lot_name"]

            self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[5]/div[2]/div[8]/div[1]/a[1]').click() #Inventory Lookup
            sleep(4)
            self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[2]/div[2]').click() #Advanced
            sleep(4)
            self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[1]/form/div[2]/div[1]/div[2]/input').send_keys(NDCOFICIAL) #Input NDC
            self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[1]/form/div[1]/div[3]/div[2]/input').send_keys(LOT_NAME) #Input Lot
            self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[2]/div[1]').click() #Search Button
            sleep(4)
            self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[2]/table/tbody/tr/td[7]/div[2]/img').click() #Select Product
            sleep(2)
            self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[1]/div/div/div/div[4]/div/div/div[1]/div/div[1]/table/thead/tr/th[1]/input').click() #Select Serial
            sleep(1)
            self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[2]/div/button[1]/span').click() #Add Selection

            sleep(4)

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click() #Save SO By Picking
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[4]/div[6]/div/div/div/input').click() #Shipped
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click() #Save in Confirm Products Quantity
        sleep(2)

        #Task - Structure that Checks the "Done" Status of the Task
        self.navegador.get(self.Link + "/utilities/background_tasks")
        sleep(2)
        while len(self.navegador.find_elements(By.XPATH, "/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr[1]/td[5]//span[text()='Done']")) != 1:
            self.navegador.refresh()
            sleep(2)

        # self.navegador.get(self.Link + "/shipments/outbound_shipments")  # Dashboard
        # sleep(2)
        #
        # self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[1]/form/div[1]/div[5]/div[2]/input').send_keys(SO_Name_SBP)  # Input SO
        # self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[2]/div[1]').click()  # Click Search
        # sleep(2)
        # self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr/td[7]/span[1]/div[5]/img').click()  # Click Delete Shipped
        # sleep(2)
        # self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/label[3]/input').click()  # Input confirmation Shipped
        # self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click()  # Confirm Cancel Shipped
        # sleep(2)

        self.End_Time_Module = datetime.now().strftime("%I:%M %p")
        self.Result = "Successfully"
        print(self.SO_Name_SBP + " Complete!")

    def outbound(self):

        self.Module = "Outbound"
        self.Start_Time_Module = datetime.now().strftime("%I:%M %p")

        #Dashboard
        sleep(2)
        self.navegador.get(self.Link) #Link

        #Create Sales Order
        #self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[5]/div[3]/div/label').click() #Create SO
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[6]/div[2]/div/label').click() #Create SO
        sleep(2)

        #Change TP
        self.navegador.find_element(By.XPATH, '/html/div[2]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[1]/form/div[1]/div[2]/div[2]/input').send_keys(self.Trading_Proc) #Search TP
        self.navegador.find_element(By.XPATH, '/html/div[2]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[2]/div').click() #Button Search o TP
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div[1]/div[2]/div/div/div/div/div/div[2]/table/tbody/tr/td[3]/div/img').click()  #Customers Select
        sleep(2)

        #Change Location
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div[3]/a').click()  # Change Location
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[1]/form/div[1]/div[2]/div[2]/input').send_keys(self.Location) #Search Location
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[2]/div').click() #Button Search
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[2]/table/tbody/tr/td[3]/div/img').click() #Select Location

        #General Tab

        # Generating PO and SO Identifications
        ph_no = []
        # A loop is used to generate 5 Numbers.
        for i in range(0, 5):
            ph_no.append(r.randint(0, 9)) #Each Number from 0 to 9
        # I take the numbers from the array and put them together
        A = ''
        for i in ph_no:
            A += str(i)
        self.SO_Name_Serial = "SO_Manual_" + A

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[6]/div[2]/input').send_keys(self.SO_Name_Serial)  #Customer Order ID
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[7]/div[2]/input').send_keys(self.SO_Name_Serial)  #Invoice
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[8]/div[2]/input').send_keys(self.SO_Name_Serial)  #SO Number
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[6]/div[4]/input').send_keys(self.SO_Name_Serial)  #Internal Reference
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[7]/div[4]/input').send_keys(self.SO_Name_Serial)  #Release
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[8]/div[4]/input').send_keys(self.SO_Name_Serial)  #Order Number
        Select(self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[9]/div[2]/div/select')).select_by_index(1)  #Select the First Option of Select Transaction Type

        #Sold By and Ship From Tab
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[2]/span').click()  #Sold By

        #Bought By and Ship To Tab
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[3]/span').click()  #Bought By

        #Line Itens Menu
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[4]/span').click()  #Items Tab
        sleep(2)

        Tb = pd.read_excel(self.file) #Add all Spreadsheet Products to the Sales Order
        for l, product in enumerate(Tb["ndc"]):
            ITEM = Tb.loc[l, "item"]
            NDCOFICIAL = Tb.loc[l, "ndc"]
            LOT_NAME = Tb.loc[l, "lot_name"]
            QUANT = str(Tb.loc[l, "quantity_lot"])

            if self.cont == 1:
                self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[4]/div[2]/div/div[2]/div/div/span').click()  # Add Product
                sleep(4)
                self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[1]/form/div[1]/div[3]/div[2]/input').send_keys(NDCOFICIAL)  # Search NDC
                self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[2]/div[1]').click()  # Search NDC Button
                sleep(4)
                self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[2]/table/tbody/tr/td[6]/div[2]/img').click()  # Click in Product
                self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[4]/div[2]/div/div[1]/div/div[2]/table/tbody/tr/td[3]/input').clear()  # Quantity Product
                sleep(2)
                self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[4]/div[2]/div/div[1]/div/div[2]/table/tbody/tr/td[3]/input').send_keys(QUANT)  # Quantity Product
                sleep(2)
            elif self.cont > 1:
                self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[4]/div[2]/div/div[2]/div/div/span').click()  # Add Product
                sleep(2)
                self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[1]/form/div[1]/div[3]/div[2]/input').send_keys(NDCOFICIAL)  # Search NDC
                self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[2]/div[1]').click()  # Search NDC Button
                sleep(2)
                self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[2]/table/tbody/tr/td[6]/div[2]/img').click()  # Click in Product
                sleep(2)
                self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[4]/div[2]/div/div[1]/div/div[2]/table/tbody/tr[' + str(ITEM) + ']/td[3]/input').clear()  # Quantity Product
                sleep(2)
                self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[4]/div[2]/div/div[1]/div/div[2]/table/tbody/tr[' + str(ITEM) + ']/td[3]/input').send_keys(QUANT)  # Quantity Product
                sleep(2)
        Tb = None

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click()  #Save SO
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div/div[2]/div[2]/p').click()  #Click in "Approve and Ship"
        sleep(2)

        #START OF THE "PICKING" SCREEN

        self.navegador.get(self.Link + "/shipments/picking") #Link
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/ul/li[1]/a').click()  #Click in Picking To Do
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[1]/form/div[1]/div[3]/div[2]/input').send_keys(self.SO_Name_Serial)  #Search SO
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[2]/div[1]').click()  #Click iin Search SO
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr/td[9]/span[1]/div[1]/img').click()  #Open to Picking SO
        sleep(5)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click()  #Pick this list
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[4]/span').click()  #Items Picked
        sleep(2)

        #Open the Table with Products
        Tb = pd.read_excel(self.file)
        for l, product in enumerate(Tb["product_name"]):
            item = Tb.loc[l, "item"]
            NAME = Tb.loc[l, "product_name"]
            NDCOFICIAL = Tb.loc[l, "ndc"]
            LOT_NAME = Tb.loc[l, "lot_name"]

            if self.Environment == "Test":
                GET_URL = 'https://api.test.tracktraceweb.com/2.0/products?identifier_us_ndc=' + NDCOFICIAL  # MAKE A GET WITH NDC
            elif self.Environment == "Production":
                GET_URL = 'https://api.tracktraceweb.com/2.0/products?identifier_us_ndc=' + NDCOFICIAL  # MAKE A GET WITH NDC

            # LINE THAT EXECUTES THE API WITH THE REQUESTS LIBRARY
            response = requests.get(GET_URL, headers=self.headers)

            if response.status_code == 200:  # IF RESPONSE IS POSITIVE
                response_dic = response.json()  # BRINGS A JSON DICTIONARY
                RA = response_dic['data'][0]  # PULLS THE UUID, WHICH IS THE FIRST TAG WITHIN THE DATA TAG
                uuid_products = (RA['uuid'])  # STORES IN A VARIABLE

                # HERE WE START THE PROCESS OF FINDING THE TRANSACTION UUID
                if self.Environment == "Test":
                    GET_URL = 'https://api.test.tracktraceweb.com/2.0/shipments?transaction_order_number=' + self.PO_Name_Serial  # MAKE A GET WITH THE PO NUMBER
                elif self.Environment == "Production":
                    GET_URL = 'https://api.tracktraceweb.com/2.0/shipments?transaction_order_number=' + self.PO_Name_Serial  # MAKE A GET WITH THE PO NUMBER

                # LINE THAT EXECUTES THE API WITH THE REQUESTS LIBRARY
                response = requests.get(GET_URL, headers=self.headers)

                if response.status_code == 200:  # IF RESPONSE IS POSITIVE
                    response_dic = response.json()  # BRINGS A JSON DICTIONARY
                    RA = response_dic['data'][0]  # PULLS THE UUID, WHICH IS THE FIRST TAG WITHIN THE DATA TAG
                    uuid_transactions = (RA['uuid'])  # STORES IN A VARIABLE

                    # HERE WE START THE PROCESS OF SEARCHING FOR THE TRANSACTION SERIALS

                    if self.Environment == "Test":
                        GET_URL = 'https://api.test.tracktraceweb.com/2.0/shipments/Inbound/' + uuid_transactions + '/serials?type=PRODUCT_LOT&serial_type=SIMPLE_SERIAL&product_uuid=' + uuid_products  # Do a GET with the PO and the UUID
                    elif self.Environment == "Production":
                        GET_URL = 'https://api.tracktraceweb.com/2.0/shipments/Inbound/' + uuid_transactions + '/serials?type=PRODUCT_LOT&serial_type=SIMPLE_SERIAL&product_uuid=' + uuid_products  # Do a GET with the PO and the UUID

                    # LINE THAT EXECUTES THE API WITH THE REQUESTS LIBRARY
                    response = requests.get(GET_URL, headers=self.headers)

                    if response.status_code == 200:  # IF THE ANSWER IS POSITIVE

                        response_dic = response.json()  # BRINGS A JSON DICTIONARY
                        for number in response_dic:  # PRINT EACH ELEMENT WITHIN JSON

                            # HERE WE START THE PROCESS OF SEARCHING FOR THE GS1_SERIAL
                            if self.Environment == "Test":
                                GET_URL = 'https://api.test.tracktraceweb.com/2.0/serial_finder?serial_type=PRODUCT_SIMPLE_SERIAL&serials=' + number + '&product_uuid=' + uuid_products + '&lot_number=' + LOT_NAME + '&result_type=ON_SCREEN&_=1705526464100'  # Do a GET with the PO and the UUID
                            elif self.Environment == "Production":
                                GET_URL = 'https://api.tracktraceweb.com/2.0/serial_finder?serial_type=PRODUCT_SIMPLE_SERIAL&serials=' + number + '&product_uuid=' + uuid_products + '&lot_number=' + LOT_NAME + '&result_type=ON_SCREEN&_=1705526464100'  # Do a GET with the PO and the UUID

                            # LINE THAT EXECUTES THE API WITH THE REQUESTS LIBRARY
                            response = requests.get(GET_URL, headers=self.headers)

                            if response.status_code == 200:  # IF THE ANSWER IS POSITIVE
                                response_dic = response.json()  # BRINGS A JSON DICTIONARY
                                GS1_Serial_Value = response_dic['results'][0]['gs1_serial']  # PULLS THE GS1_SERIAL

                                self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[4]/div[2]/div[3]/input').send_keys(GS1_Serial_Value)  # Input GS1 Serial
                                self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[4]/div[2]/div[4]/a').click()  # Add/Remove Item
                                sleep(2)
                                self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[4]/div[2]/div[3]/input').clear() # Clearnning Input
                            else:
                                print(response.status_code)

        Tb = None

        #Picking Lot
        # self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[1]/div[6]/div[2]/a').click()  #Choose Lot
        # sleep(2)
        # self.navegador.find_element(By.XPATH, '/html/div[6]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[1]/form/div[1]/div[4]/div[2]/input').send_keys("TP01PD01LLB01") #Lot Picking
        # self.navegador.find_element(By.XPATH, '/html/div[6]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[2]/div[1]').click()  #Button Search Lot
        # sleep(5)
        # self.navegador.find_element(By.XPATH, '/html/div[6]/div[1]/div[2]/div/div/div/div/div/div[2]/table/tbody/tr/td[5]/div/img').click()  #Select Lot
        # sleep(2)
        # self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[2]/div[3]/div[1]/div[2]/input').send_keys("1")  #Serial Picking
        # self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[2]/div[3]/div[1]/div[3]/a').click()  #Select o Pick do Item
        # sleep(2)
        # self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button/span').click()  #Close Picking
        # sleep(2)
        # self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[3]/span').click()  #OK in View Picking List
        # sleep(2)

        #Complete Picking
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[3]/span').click()  #OK in View Picking List
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div/div/input').click()  # Select Radio Shipped
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[7]/div/div/div/input').click()  # Select Radio Shipped
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click()  #OK in Complete Picking
        sleep(2)

        # Task - Structure that Checks the "Done" Status of the Task
        self.navegador.get(self.Link + "/utilities/background_tasks")
        sleep(2)
        while len(self.navegador.find_elements(By.XPATH,"/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr[1]/td[5]//span[text()='Done']")) != 1:
            self.navegador.refresh()
            sleep(2)

        self.navegador.get(self.Link + "/shipments/outbound_shipments")  # Dashboard
        sleep(2)

        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[1]/form/div[1]/div[5]/div[2]/input').send_keys(self.SO_Name_Serial)  # Input SO
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[2]/div[1]').click()  # Click Search
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr/td[7]/span[1]/div[5]/img').click()  # Click Delete Shipped
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/label[3]/input').click()  # Input confirmation Shipped
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click()  # Confirm Cancel Shipped
        sleep(2)

        self.End_Time_Module = datetime.now().strftime("%I:%M %p")
        self.Result = "Successfully"
        print(self.SO_Name_Serial + " Complete!")

    def containers(self):

        self.Module = "Containers"
        self.Start_Time_Module = datetime.now().strftime("%I:%M %p")

        self.navegador.get(self.Link + "/company_mgt/")  # Company Management Link
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div[1]/ul/li[1]/a').click()  # Select Company Settings
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[4]/span').click()  # Select Products Tab
        sleep(2)
        input_element = self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[4]/div[5]/div/div[2]/div[2]/div/div/input')
        GS1_Company_Prefix = input_element.get_attribute("value")

        # Create Container
        self.navegador.get(self.Link + "/containers") #Containers Management Link
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/ul/li[3]/a/span').click()  #Select Create New Container
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[2]/div[3]/div/input').click()  #Change GS1 URN

        ph_no = []
        # A loop is used to generate 10 Numbers.
        for i in range(0, 12):
            ph_no.append(r.randint(0, 9))  # Each Number from 0 to 9
        # I take the numbers from the array and put them together
        self.Container_URN = ''
        for i in ph_no:
            self.Container_URN += str(i)

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[2]/div[4]/input').send_keys("urn:epc:id:sscc:" + GS1_Company_Prefix + ".1" + self.Container_URN)  #Input GS1URN Container
        Select(self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[8]/div[2]/div/select')).select_by_visible_text(self.Location)
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click()  #Save New Container
        sleep(8)

        #Task - Structure that Checks the "Done" Status of the Task
        self.navegador.get(self.Link + "/utilities/background_tasks")
        sleep(2)
        while len(self.navegador.find_elements(By.XPATH, "/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr[1]/td[5]//span[text()='Done']")) != 1:
            self.navegador.refresh()
            sleep(2)

        #Edita os Produtos do Container
        self.navegador.get(self.Link + "/containers")  #Containers Management Link
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/ul/li[1]/a/label').click()  #List Container
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[1]/form/div[1]/div[2]/div[2]/input').send_keys("urn:epc:id:sscc:" + GS1_Company_Prefix + ".1" + self.Container_URN) #Search Containers
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[2]/div[1]').click()  #Click in Search Container
        sleep(4)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr/td[9]/span[1]/div[2]/img').click()  #Edit Container
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[2]').click()  #Items Tab
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[1]/div/div[2]/div[3]/span[1]').click()  # Click in Add with GS1
        sleep(2)

        # Open the Table with Products
        Tb = pd.read_excel(self.file)
        for l, product in enumerate(Tb["product_name"]):
            item = Tb.loc[l, "item"]
            NDCOFICIAL = Tb.loc[l, "ndc"]
            LOT_NAME = Tb.loc[l, "lot_name"]

            if self.Environment == "Test":
                GET_URL = 'https://api.test.tracktraceweb.com/2.0/products?identifier_us_ndc=' + NDCOFICIAL  # MAKE A GET WITH NDC
            elif self.Environment == "Production":
                GET_URL = 'https://api.tracktraceweb.com/2.0/products?identifier_us_ndc=' + NDCOFICIAL  # MAKE A GET WITH NDC

            # LINE THAT EXECUTES THE API WITH THE REQUESTS LIBRARY
            response = requests.get(GET_URL, headers=self.headers)

            if response.status_code == 200:  # IF RESPONSE IS POSITIVE
                response_dic = response.json()  # BRINGS A JSON DICTIONARY
                RA = response_dic['data'][0]  # PULLS THE UUID, WHICH IS THE FIRST TAG WITHIN THE DATA TAG
                uuid_products = (RA['uuid'])  # STORES IN A VARIABLE

                # HERE WE START THE PROCESS OF FINDING THE TRANSACTION UUID
                if self.Environment == "Test":
                    GET_URL = 'https://api.test.tracktraceweb.com/2.0/shipments?transaction_order_number=' + self.PO_Name_Serial  # MAKE A GET WITH THE SO NUMBER
                elif self.Environment == "Production":
                    GET_URL = 'https://api.tracktraceweb.com/2.0/shipments?transaction_order_number=' + self.PO_Name_Serial  # MAKE A GET WITH THE SO NUMBER

                # LINE THAT EXECUTES THE API WITH THE REQUESTS LIBRARY
                response = requests.get(GET_URL, headers=self.headers)

                if response.status_code == 200:  # IF RESPONSE IS POSITIVE
                    response_dic = response.json()  # BRINGS A JSON DICTIONARY
                    RA = response_dic['data'][0]  # PULLS THE UUID, WHICH IS THE FIRST TAG WITHIN THE DATA TAG
                    uuid_transactions = (RA['uuid'])  # STORES IN A VARIABLE

                    # HERE WE START THE PROCESS OF SEARCHING FOR THE TRANSACTION SERIALS
                    if self.Environment == "Test":
                        GET_URL = 'https://api.test.tracktraceweb.com/2.0/shipments/Inbound/' + uuid_transactions + '/serials?type=PRODUCT_LOT&serial_type=SIMPLE_SERIAL&product_uuid=' + uuid_products  # Do a GET with the PO and the UUID
                    elif self.Environment == "Production":
                        GET_URL = 'https://api.tracktraceweb.com/2.0/shipments/Inbound/' + uuid_transactions + '/serials?type=PRODUCT_LOT&serial_type=SIMPLE_SERIAL&product_uuid=' + uuid_products  # Do a GET with the PO and the UUID

                    # LINE THAT EXECUTES THE API WITH THE REQUESTS LIBRARY
                    response = requests.get(GET_URL, headers=self.headers)

                    if response.status_code == 200:  # IF THE ANSWER IS POSITIVE

                        response_dic = response.json()  # BRINGS A JSON DICTIONARY
                        for number in response_dic:  # PRINT EACH ELEMENT WITHIN JSON

                            # HERE WE START THE PROCESS OF SEARCHING FOR THE GS1_SERIAL
                            if self.Environment == "Test":
                                GET_URL = 'https://api.test.tracktraceweb.com/2.0/serial_finder?serial_type=PRODUCT_SIMPLE_SERIAL&serials=' + number + '&product_uuid=' + uuid_products + '&lot_number=' + LOT_NAME + '&result_type=ON_SCREEN&_=1705526464100'  # Do a GET with the PO and the UUID
                            elif self.Environment == "Production":
                                GET_URL = 'https://api.tracktraceweb.com/2.0/serial_finder?serial_type=PRODUCT_SIMPLE_SERIAL&serials=' + number + '&product_uuid=' + uuid_products + '&lot_number=' + LOT_NAME + '&result_type=ON_SCREEN&_=1705526464100'  # Do a GET with the PO and the UUID

                            # LINE THAT EXECUTES THE API WITH THE REQUESTS LIBRARY
                            response = requests.get(GET_URL, headers=self.headers)

                            if response.status_code == 200:  # IF THE ANSWER IS POSITIVE
                                response_dic = response.json()  # BRINGS A JSON DICTIONARY
                                GS1_Serial_Value = response_dic['results'][0]['gs1_serial']  # PULLS THE GS1_SERIAL

                                self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[2]/div/textarea').send_keys(GS1_Serial_Value + "\n")  # Paste the GS1_Serial Input
                                sleep(2)
                            else:
                                print(response.status_code)

        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click()  # Click OK in Add GS1 Serialized
        sleep(6)

        # Task - Structure that Checks the "Done" Status of the Task
        self.navegador.get(self.Link + "/utilities/background_tasks")
        sleep(2)
        while len(self.navegador.find_elements(By.XPATH, "/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr[1]/td[5]//span[text()='Done']")) != 1:
            self.navegador.refresh()
            sleep(2)

        #Deletion of Container Products
        self.navegador.get(self.Link + "/containers")  # Link do Containers Management
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/ul/li[1]/a/label').click()  # List Container
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[1]/form/div[1]/div[2]/div[2]/input').send_keys("urn:epc:id:sscc:" + GS1_Company_Prefix + ".1" + self.Container_URN) #Search for the created Container
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[2]/div[1]').click()  #Click on Search Container
        sleep(4)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr/td[9]/span[1]/div[2]/img').click()  # Edit Container
        sleep(4)

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[2]').click()  # Items Tab
        sleep(4)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[1]/div/div[1]/div[1]/div/div[1]/table/thead/tr/th[1]/input').click()  #Select All Item
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[1]/div/div[2]/div[1]/span').click()  #Remove All Item
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click()  #Ok in Remove All Item
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[4]/div[1]/div/input').click()  #Select "Move Items to Storage Area"
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click()  #Click in Remove Item
        sleep(6)

        # Task - Structure that Checks the "Done" Status of the Task
        self.navegador.get(self.Link + "/utilities/background_tasks")
        sleep(2)
        while len(self.navegador.find_elements(By.XPATH, "/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr[1]/td[5]//span[text()='Done']")) != 1:
            self.navegador.refresh()
            sleep(2)

        #Deletion of Container
        self.navegador.get(self.Link + "/containers")  # Link do Containers Management
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/ul/li[1]/a/label').click()  # List Container
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[1]/form/div[1]/div[2]/div[2]/input').send_keys("urn:epc:id:sscc:" + GS1_Company_Prefix + ".1" + self.Container_URN)  # Search for the created Container
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[2]/div[1]').click()  # Click on Search Container
        sleep(4)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr/td[9]/span[1]/div[2]/img').click()  # Click on Container
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[2]/span').click()  #Click Delete Container
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click()  #Click OK in Delete Container
        sleep(2)

        #Task - Structure that Checks the "Done" Status of the Task
        self.navegador.get(self.Link + "/utilities/background_tasks")
        sleep(2)
        while len(self.navegador.find_elements(By.XPATH, "/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr[1]/td[5]//span[text()='Done']")) != 1:
            self.navegador.refresh()
            sleep(2)

        self.End_Time_Module = datetime.now().strftime("%I:%M %p")
        self.Result = "Successfully"
        print("Container Complete!")

    def quarantine(self):

        self.Module = "Quarantine"
        self.Start_Time_Module = datetime.now().strftime("%I:%M %p")

        self.navegador.get(self.Link + "/quarantine") #Link Quarentine
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/ul/li[1]/a').click()  #Select Quarantine Items
        sleep(2)

        #Change Location
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[1]/div[3]/a').click()  # Change Location
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[1]/form/div[1]/div[2]/div[2]/input').send_keys(self.Location) #Search Location
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[2]/div').click() #Button Search
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[2]/table/tbody/tr/td[3]/div/img').click() #Select Location
        sleep(2)

        #General Tab
        ph_no = []
        # A loop is used to generate 5 Numbers.
        for i in range(0, 4):
            ph_no.append(r.randint(0, 9))  # Each Number from 0 to 9
        # I take the numbers from the array and put them together
        QR = ''
        for i in ph_no:
            QR += str(i)
        self.Quarantine_Ref = "Quarantine_" + QR
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[4]/div[2]/input').send_keys(self.Quarantine_Ref)  # TP Name Field
        Select(self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[3]/div[2]/div[1]/div[2]/div/select')).select_by_index(3)  # Select Reason
        sleep(2)

        #Items Tab
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[2]/span').click()  #Aba Items
        sleep(2)

        Tb = pd.read_excel(self.file)  # Adds the Aggregation of all Purchase Order Products
        for l, product in enumerate(Tb["product_name"]):
            NDCOFICIAL = Tb.loc[l, "ndc"]
            LOT_NAME = Tb.loc[l, "lot_name"]

            self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[2]/div/div[2]/div[4]/span').click()  #Add with Item
            sleep(2)
            self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[1]/form/div[1]/div[3]/div[2]/input').send_keys(NDCOFICIAL)
            self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[2]/div[1]').click()  #Pesquisar
            sleep(2)
            self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[2]/table/tbody/tr[1]/td[6]/div[2]/img').click()  #Click Product
            sleep(2)
            self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[3]/div[2]/input').send_keys(LOT_NAME)
            self.navegador.find_element(By.XPATH, '//*[@id="serach_btn"]/span').click()  #Search
            sleep(2)
            self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[4]/div/div/div[1]/div/div[2]/table/tbody/tr/td[4]/span[1]/div/img').click()  #Seleciona Lot
            sleep(2)
            self.navegador.find_element(By.XPATH, '//*[@id="destructions_form__list__serials_select_all"]').click()  #Seleciona Serial
            sleep(2)
            self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[2]/div/button[1]/span').click() #Add Selection
            sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click()  #OK New Quarentine
        sleep(6)

        # Task - Structure that Checks the "Done" Status of the Task
        self.navegador.get(self.Link + "/utilities/background_tasks")
        sleep(2)
        while len(self.navegador.find_elements(By.XPATH, "/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr[1]/td[5]//span[text()='Done']")) != 1:
            self.navegador.refresh()
            sleep(2)

        #Unquarantine
        self.navegador.get(self.Link + "/quarantine") #Link Quarantine
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/ul/li[2]/a').click()  #Quarantine Items
        sleep(2)

        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[1]/form/div[1]/div[5]/div[2]/input').send_keys(self.Quarantine_Ref)  #Input Reference
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[2]/div[1]').click()  #Search Quarantine
        sleep(4)

        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr/td[7]/span[1]/div[2]/img').click()  #Un-Quarantine
        sleep(6)
        Select(self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[4]/div[2]/div[1]/div[2]/div/select')).select_by_index(1)  #Select Reason
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click()  #Ok em Un-Quarentine
        sleep(6)

        # Task - Structure that Checks the "Done" Status of the Task
        self.navegador.get(self.Link + "/utilities/background_tasks")
        sleep(2)
        while len(self.navegador.find_elements(By.XPATH, "/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr[1]/td[5]//span[text()='Done']")) != 1:
            self.navegador.refresh()
            sleep(2)

        self.End_Time_Module = datetime.now().strftime("%I:%M %p")
        self.Result = "Successfully"
        print("Quarantine Complete!")

    def RMA(self):

        self.Module = "RMA"
        self.Start_Time_Module = datetime.now().strftime("%I:%M %p")

        #Create RMA
        self.navegador.get(self.Link + "/return_manager/")  # Link
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div[1]/ul/li[1]/a').click()  #Click Create RMA
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[1]/div[3]/div[4]/div[1]/div/input').click()  #Click RMA Number

        ph_no = []
        # A loop is used to generate 5 Numbers.
        for i in range(0, 5):
            ph_no.append(r.randint(0, 9))  # Each Number from 0 to 9
        # I take the numbers from the array and put them together
        A = ''
        for i in ph_no:
            A += str(i)
        RMA = "RMA_" + A

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[1]/div[3]/div[4]/div[2]/input').send_keys(RMA)  # RMA Number
        Select(self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[1]/div[7]/div[2]/div[2]/div/select')).select_by_index(1)  # Select the First ComboBox Option

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[1]/div[8]/div[3]/a').click()  # Click Select Customers
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[1]/form/div[1]/div[2]/div[2]/input').send_keys(self.Trading_Proc) #Input Trading
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[2]/div').click() #Search Button
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[2]/table/tbody/tr/td[3]/div/img').click()  # Select Customers
        sleep(2)

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div[2]/div[4]/a').click()  # Click Select Shipment
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[1]/form/div[1]/div[4]/div[2]/input').send_keys(self.SO_Name_SBP) #Input SO
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[2]/div[1]').click() #Search Button
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[2]/table/tbody/tr/td[6]/div[2]/img').click()  #Select SO
        sleep(2)

        Select(self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[1]/div[7]/div[2]/div[2]/div/select')).select_by_index(1)  # Select Product Select
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[3]/span').click()  #Change Items Tab
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[3]/div[2]/div[2]/div/div[2]/span').click()  #Click Return All
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click()  #OK in Create RMA
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click()  #Yes in Approve RMA
        sleep(2)

        #Proceed to a Return
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div[2]/ul/li[1]/a').click()  #Link Proceed to a return
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[5]/div/div/input').click()  #Select Existing RMA
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[6]/div[3]').click()  #Select RMA
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[1]/form/div[1]/div[2]/div[2]/input').send_keys(RMA) #Input RMA ID
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[2]/div').click() #Search Button
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[2]/table/tbody/tr/td[6]/div[2]/img').click()  #Select RMA
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click()  #OK in Proceed Return
        sleep(5)

        ph_no = []
        # A loop is used to generate 5 Numbers.
        for i in range(0, 5):
            ph_no.append(r.randint(0, 9))  # Each Number from 0 to 9
        # I take the numbers from the array and put them together
        A = ''
        for i in ph_no:
            A += str(i)
        RTN = "REF#" + A

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[1]/div[3]/div[2]/input').send_keys(RTN)  #Ref Tracking Number
        Select(self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[1]/div[6]/div[2]/div/select')).select_by_index(1)  # Select the First ComboBox Option
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[1]/div[7]/div[2]/div/input').click()  #Remove VRS

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[4]/span').click()  #Change to Serial Entry Tab
        sleep(2)
        Select(self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[4]/div[2]/div[2]/div[2]/div/select')).select_by_index(2)  #Simple Serial

        Tb = pd.read_excel(self.file)
        for l, product in enumerate(Tb["product_name"]):
            NDCOFICIAL = Tb.loc[l, "ndc"]
            LOT_NAME = Tb.loc[l, "lot_name"]
            EXP = Tb.loc[l, "exp"]

            self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[4]/div[2]/div[4]/div[3]/a').click()  # Select Product
            sleep(2)
            self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[1]/form/div[1]/div[3]/div[2]/input').send_keys(NDCOFICIAL)  # Product
            self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[2]/div[1]').click()  # Search
            sleep(2)
            self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[2]/table/tbody/tr[1]/td[6]/div[2]/img').click()  # Select Product
            sleep(2)

            self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[4]/div[2]/div[6]/div[2]/input').send_keys(LOT_NAME)  # Lot
            self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[4]/div[2]/div[6]/div[4]/input').send_keys(EXP)  # Exp

            # Processo de RMA dos Seriais

            # AQUI INICIAMOS O PROCESSO DE BUSCAR O UUID DO PRODUTO

            if self.Environment == "Test":
                GET_URL = 'https://api.test.tracktraceweb.com/2.0/products?identifier_us_ndc=' + NDCOFICIAL  # FAZ UM GET COM O NDC
            elif self.Environment == "Production":
                GET_URL = 'https://api.tracktraceweb.com/2.0/products?identifier_us_ndc=' + NDCOFICIAL  # FAZ UM GET COM O NDC

            # LINHA QUE EXECULTA A API COM A BIBLIOTECA REQUESTS
            response = requests.get(GET_URL, headers=self.headers)

            if response.status_code == 200:  # SE A RESPOSTA FOR POSITIVO
                response_dic = response.json()  # TRAZ UM DICIONARIO JSON
                RA = response_dic['data'][0]  # PUXA O UUID, QUE É A PRIMEIRA TAG DENTRO DA TAG DATA
                uuid_products = (RA['uuid'])  # ARMAZENA EM UMA VARIAVEL

                # AQUI INICIAMOS O PROCESSO DE BUSCAR O UUID DA TRANSAÇÃO
                if self.Environment == "Test":
                    GET_URL = 'https://api.test.tracktraceweb.com/2.0/shipments?transaction_order_number=' + self.SO_Name_SBP  # FAZ UM GET COM O NUMERO DA SO
                elif self.Environment == "Production":
                    GET_URL = 'https://api.tracktraceweb.com/2.0/shipments?transaction_order_number=' + self.SO_Name_SBP  # FAZ UM GET COM O NUMERO DA SO

                # LINHA QUE EXECULTA A API COM A BIBLIOTECA REQUESTS
                response = requests.get(GET_URL, headers=self.headers)

                if response.status_code == 200:  # SE A RESPOSTA FOR POSITIVO
                    response_dic = response.json()  # TRAZ UM DICIONARIO JSON
                    RA = response_dic['data'][0]  # PUXA O UUID, QUE É A PRIMEIRA TAG DENTRO DA TAG DATA
                    uuid_transactions = (RA['uuid'])  # ARMAZENA EM UMA VARIAVEL

                    # AQUI INICIAMOS O PROCESSO DE BUSCAR OS SERIAIS DA TRANSACTION
                    if self.Environment == "Test":
                        GET_URL = 'https://api.test.tracktraceweb.com/2.0/shipments/Outbound/' + uuid_transactions + '/serials?type=PRODUCT_LOT&serial_type=SIMPLE_SERIAL&product_uuid=' + uuid_products  # FAZ UM GET COM O NUMERO DA PO E O UUID DO PRODUTO
                    elif self.Environment == "Production":
                        GET_URL = 'https://api.tracktraceweb.com/2.0/shipments/Outbound/' + uuid_transactions + '/serials?type=PRODUCT_LOT&serial_type=SIMPLE_SERIAL&product_uuid=' + uuid_products  # FAZ UM GET COM O NUMERO DA PO E O UUID DO PRODUTO

                    # LINHA QUE EXECULTA A API COM A BIBLIOTECA REQUESTS
                    response = requests.get(GET_URL, headers=self.headers)

                    if response.status_code == 200:  # SE A RESPOSTA FOR POSITIVO

                        response_dic = response.json()  # TRAZ UM DICIONARIO JSON
                        for numero in response_dic:  # IMPRIME CADA ELEMENTO DENTRO DO JSON
                            self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[4]/div[2]/div[5]/div[2]/input').send_keys(numero) #Coloca todos os Numeros de Seriais no RMA
                            sleep(1)
                            self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[4]/div[2]/div[5]/div[3]/a').click()  #Clica em Add Product
                            self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[4]/div[2]/div[5]/div[2]/input').clear()

                    else:
                        print(response.status_code)
                else:
                    print(response.status_code)
            else:
                print(response.status_code)

            self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[4]/div[2]/div[6]/div[2]/input').clear()  # Limpa Input Lot
            self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[4]/div[2]/div[6]/div[4]/input').clear()  # Limpa o Input de Validade

        sleep(5)

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click() #Save RMA
        sleep(2)

        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[4]/div[2]/div/input').click()  # Close RMA
        sleep(2)
        Select(self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[7]/div[3]/div[2]/div/select')).select_by_visible_text(self.Location)  # Select Location
        sleep(1)
        Select(self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[7]/div[4]/div[2]/div/select')).select_by_index(1)  # Select Storage Area
        sleep(1)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click()  # Ok in Return Merchandise
        sleep(2)

        #Task - Structure that Checks the "Done" Status of the Task
        self.navegador.get(self.Link + "/utilities/background_tasks")
        sleep(2)
        while len(self.navegador.find_elements(By.XPATH, "/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr[1]/td[5]//span[text()='Done']")) != 1:
            self.navegador.refresh()
            sleep(2)

        self.End_Time_Module = datetime.now().strftime("%I:%M %p")
        self.Result = "Successfully"
        print(RMA + " Complete!")

    def com_pack(self):

        self.Module = "Commission and Packaging"
        self.Start_Time_Module = datetime.now().strftime("%I:%M %p")

        #START OF THE "COMISSION AND PACKAGING/UNPACKING" PROCESS

        #START MANAGEMENT PRODUCTS

        # Generating Identifications Separately

        # GENERATES THE PREFIX, THE TWO NUMBERS THAT PRECED THE FIRST PART OF THE NDC IN THE COMPANY PREFIX.
        ph_no = []
        # A loop is used to generate 2 Numbers.
        for i in range(0, 2):
            ph_no.append(r.randint(0, 9))  # Each Number from 0 to 9
        # I take the numbers from the array and put them together
        W1 = ''
        for i in ph_no:
            W1 += str(i)

        # GENERATES FIRST PART OF THE NDC, COMPOSED OF 05 DIGITS.
        ph_no = []
        # A loop is used to generate 5 Numbers.
        for i in range(0, 5):
            ph_no.append(r.randint(0, 9))  # Each Number from 0 to 9
        # I take the numbers from the array and put them together
        X1 = ''
        for i in ph_no:
            X1 += str(i)

        # GENERATES SECOND PART OF THE NDC, COMPOSED OF 03 DIGITS.
        ph_no = []
        # A loop is used to generate 3 Numbers.
        for i in range(0, 3):
            ph_no.append(r.randint(0, 9))  # Each Number from 0 to 9
        # I take the numbers from the array and put them together
        Y1 = ''
        for i in ph_no:
            Y1 += str(i)

        # GENERATES SECOND PART OF THE NDC, COMPOSED OF 03 DIGITS.
        ph_no = []
        # A loop is used to generate 5 Numbers.
        for i in range(0, 2):
            ph_no.append(r.randint(0, 9))  # Each Number from 0 to 9
        # I take the numbers from the array and put them together
        Z1 = ''
        for i in ph_no:
            Z1 += str(i)

        # GENERATES PRODUCT NUMBER
        W = W1
        X = X1
        Y = Y1
        Z = Z1
        GS1PREFIX_CHILD = W + X
        GS1ID_CHILD = "0" + Y + Z
        NDCOFICIAL_CHILD = X + "-" + Y + "-" + Z

        #Dashboard
        sleep(2)
        self.navegador.get(self.Link) #Link

        # Open Products Screen
        self.navegador.get(self.Link + "/products/")  # Product Manager Link
        sleep(2)

        # Add Child Products
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[2]/div/div[3]/span').click()  # Add Button
        sleep(2)
        self.navegador.find_element(By.XPATH, '//*[@id="TT_UTILS_UI_FORM_UUID__1_name"]').send_keys("Child Product Bottle")  # Name Product
        Select(self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[3]/div[3]/div/select')).select_by_index(2)  # Select the Bottle
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[2]/span').click()  # Identifier Tab
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[1]/div[2]/input').send_keys("9999999999")  # SKU
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[2]/div[2]/input').send_keys("888888888888")  # UPC
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[3]/div[2]/input').send_keys(GS1PREFIX_CHILD)  # GS1 Company Prefix
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[4]/div[2]/input').send_keys(GS1ID_CHILD)  # GS1 ID
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[5]/div[2]/div[2]/div/div/span').click()  # Add Identifiers
        sleep(2)
        Select(self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div/select')).select_by_index(0)  # Select the First ComboBox Option
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[2]/div[2]/input').send_keys(NDCOFICIAL_CHILD)  # NDC
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click()  # Add in Add Identifier
        sleep(2)

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[5]/span').click()  # Misc Tab
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[5]/div[6]/div/div/input').click()  # Deactivate Leaf

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click()  # Add in Produto Forms

        # Generating Identifications Separately

        # GENERATES THE PREFIX, THE TWO NUMBERS THAT PRECED THE FIRST PART OF THE NDC IN THE COMPANY PREFIX.
        ph_no = []
        # A loop is used to generate 2 Numbers.
        for i in range(0, 2):
            ph_no.append(r.randint(0, 9))  # Each Number from 0 to 9
        # I take the numbers from the array and put them together
        W1 = ''
        for i in ph_no:
            W1 += str(i)

        # GENERATES FIRST PART OF THE NDC, COMPOSED OF 05 DIGITS.
        ph_no = []
        # A loop is used to generate 5 Numbers.
        for i in range(0, 5):
            ph_no.append(r.randint(0, 9))  # Each Number from 0 to 9
        # I take the numbers from the array and put them together
        X1 = ''
        for i in ph_no:
            X1 += str(i)

        # GENERATES SECOND PART OF THE NDC, COMPOSED OF 03 DIGITS.
        ph_no = []
        # A loop is used to generate 3 Numbers.
        for i in range(0, 3):
            ph_no.append(r.randint(0, 9))  # Each Number from 0 to 9
        # I take the numbers from the array and put them together
        Y1 = ''
        for i in ph_no:
            Y1 += str(i)

        # GENERATES SECOND PART OF THE NDC, COMPOSED OF 03 DIGITS.
        ph_no = []
        # A loop is used to generate 5 Numbers.
        for i in range(0, 2):
            ph_no.append(r.randint(0, 9))  # Each Number from 0 to 9
        # I take the numbers from the array and put them together
        Z1 = ''
        for i in ph_no:
            Z1 += str(i)

        # GENERATES PRODUCT NUMBER
        W = W1
        X = X1
        Y = Y1
        Z = Z1
        GS1PREFIX_FATHER = W + X
        GS1ID_FATHER = "0" + Y + Z
        NDCOFICIAL_FATHER = X + "-" + Y + "-" + Z

        sleep(2)

        # Add Father Products
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[2]/div/div[3]/span').click()  # Add Button
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div[3]/div[2]/input').send_keys("Father Product Case")  # Name Product
        Select(self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[3]/div[3]/div/select')).select_by_index(4)  # Select the Case
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[2]/span').click()  # Identifier Tab
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[1]/div[2]/input').send_keys("9999999999")  # SKU
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[2]/div[2]/input').send_keys("888888888888")  # UPC
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[3]/div[2]/input').send_keys(GS1PREFIX_FATHER)  # GS1 Company Prefix
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[4]/div[2]/input').send_keys(GS1ID_FATHER)  # GS1 ID
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[5]/div[2]/div[2]/div/div/span').click()  # Add Identifiers
        sleep(2)
        Select(self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div/select')).select_by_index(0)  # Select the First ComboBox Option
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[2]/div[2]/input').send_keys(NDCOFICIAL_FATHER)  # NDC
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click()  # Add in Add Identifier
        sleep(2)

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[4]/span').click()  # Aggregation Tab
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[4]/div[2]/div[2]/a').click()  # Add Product
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[1]/form/div[1]/div[2]/div[2]/input').send_keys(NDCOFICIAL_CHILD)  # NDC
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[2]/div[1]').click()  # Search Child
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[2]/table/tbody/tr[1]/td[6]/div[2]/img').click()  #Select Child
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[5]/div[2]/input').send_keys("2") #Quantity Child Product
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click()  #Ok in Quantity
        sleep(2)

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[5]/span').click()  # Misc Tab
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[5]/div[6]/div/div/input').click()  # Deactivate Leaf

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click()  # Add in Product Forms

        #START OF THE "COMMISSION" SCREEN

        self.navegador.get(self.Link + "/company_mgt/locations") #Link
        sleep(2)

        #Change Location
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[1]/form/div[1]/div[2]/div[2]/input').send_keys(self.Location) #Search Location
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[2]/div[1]').click() #Button Search
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr/td[8]/span[1]/div[2]/img').click()  #Edit Location
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[7]/span').click()  #Change Tab
        sleep(2)

        #Task - Structure that Checks Whether or Not There Are Products in Manufacturing.
        if len(self.navegador.find_elements(By.XPATH, "/html/div[2]/div/div[2]/div[1]/div/div/div/div[7]/div/div[1]/div/div[2]/table/tbody/tr/td[5]/span[1]/div[3]/img")) == 1:
            self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[7]/div/div[1]/div/div[2]/table/tbody/tr/td[5]/span[1]/div[3]/img').click()  # Delete Product
            sleep(2)
            self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click()  # Confirm Delete Product
            sleep(2)

        #Add Child in Manufacturing
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[7]/div/div[2]/div/div/span').click()  #Add Manufacturing Button
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[1]/form/div[1]/div[2]/div[2]/input').send_keys(NDCOFICIAL_CHILD)  #Edit Location
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[2]/div[1]').click()  #Search Button
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[2]/table/tbody/tr[1]/td[6]/div[2]/img').click()  #Select Product
        sleep(2)
        self.navegador.find_element(By.XPATH, '//html/div[4]/div/div[2]/div[2]/div/button[1]/span').click()  #Add Product
        sleep(2)

        #Add Father in Manufacturing
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[7]/div/div[2]/div/div/span').click()  #Add Manufacturing Button
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[1]/form/div[1]/div[2]/div[2]/input').send_keys(NDCOFICIAL_FATHER)  #Edit Location
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[2]/div[1]').click()  #Search Button
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[2]/table/tbody/tr[1]/td[6]/div[2]/img').click()  #Select Product
        sleep(2)
        self.navegador.find_element(By.XPATH, '//html/div[4]/div/div[2]/div[2]/div/button[1]/span').click()  #Add Product
        sleep(2)

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click()  #Save

        # Add Serialized Child
        self.navegador.get(self.Link + "/products/manufacturer") #Link Lot Creation and Serial Generation
        sleep(2)

        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[2]/div/div[2]/span').click()  #Add Serialied Lot
        sleep(2)
        Select(self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[1]/div[2]/div[2]/div/select')).select_by_visible_text("Child Product Bottle") #Select Product Select
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div[3]/div[2]/input').send_keys("CHILDLOT")  #Lot Name
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div[5]/div[2]/input').send_keys("08-08-2025")  # Input Exp
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[2]/span').click() #Save Lot
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[1]/form/div[1]/div[3]/div[2]/input').send_keys("CHILDLOT")  #Search Lot Name
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[2]/div[1]').click() #Search Button
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr/td[9]/span[1]/div[2]/img').click() #Edit Lot
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[2]/span').click() #Serials Lot
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div/div[2]/div/div/span').click() #New Serials Request
        sleep(4)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[2]/div[2]/input').send_keys("2") #Quantity to generate
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click() #Add Request
        sleep(4)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click() #Dimiss Button
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[2]/span').click() #Ok in Edit Lot
        sleep(2)

        #Add Serialized Father
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[2]/div/div[2]/span').click()  #Add Serialied Lot
        sleep(2)
        Select(self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[1]/div[2]/div[2]/div/select')).select_by_visible_text("Father Product Case") #Select Product Select
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div[3]/div[2]/input').send_keys("FATHERLOT")  #Lot Name
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div[5]/div[2]/input').send_keys("08-08-2025")  # Input Exp
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[2]/span').click() #Save Lot
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[1]/form/div[1]/div[3]/div[2]/input').clear() #Clear Field
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[1]/form/div[1]/div[3]/div[2]/input').send_keys("FATHERLOT")  #Search Lot Name
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[2]/div[1]').click() #Search Button
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr/td[9]/span[1]/div[2]/img').click() #Edit Lot
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[2]/span').click() #Serials Lot
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div/div[2]/div/div/span').click() #New Serials Request
        sleep(4)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[2]/div[2]/input').send_keys("1") #Quantity to generate
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click() #Add Request
        sleep(4)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click() #Dimiss Button
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[2]/span').click() #Ok in Edit Lot
        sleep(2)

        self.navegador.get(self.Link + "/commision_serials") #Comission Link
        sleep(2)

        #COMMISSION CHILD
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[1]/form/div[1]/div[5]/div[2]/input').send_keys("CHILDLOT") #Input Products
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[2]/div[1]').click() #Search
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr/td[9]/span[1]/div[2]/img').click() #Confirm Lot
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[5]/div/a').click() #Select Serials
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[4]/div/div/div[1]/div/div[1]/table/thead/tr/th[1]/input').click() #Select All Serials
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click() #Select Serials Numbers Button
        sleep(2)
        Select(self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[12]/div[2]/div/select')).select_by_index(1) #Select Storage Area
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click() #Save Lot
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click() #Dimiss Button

        #COMMISSION FATHER
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[1]/form/div[1]/div[5]/div[2]/input').clear()
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[1]/form/div[1]/div[5]/div[2]/input').send_keys("FATHERLOT") #Input Products
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[2]/div[1]').click() #Search
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr/td[9]/span[1]/div[2]/img').click() #Confirm Lot
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[5]/div/a').click() #Select Serials
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[4]/div/div/div[1]/div/div[1]/table/thead/tr/th[1]/input').click() #Select All Serials
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click() #Select Serials Numbers Button
        sleep(2)
        Select(self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[12]/div[2]/div/select')).select_by_index(1) #Select Storage Area
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click() #Save Lot
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click() #Dimiss Button

        #Task - Structure that Checks the "Done" Status of the Task
        self.navegador.get(self.Link + "/utilities/background_tasks")
        sleep(2)
        while len(self.navegador.find_elements(By.XPATH, "/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr[1]/td[5]//span[text()='Done']")) != 1:
            self.navegador.refresh()
            sleep(2)

        # START OF THE "PACKAGING" SCREEN
        self.navegador.get(self.Link + "/packaging/") #Packaging Link
        sleep(4)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[1]/form/div[1]/div[5]/div[2]/input').send_keys("FATHERLOT") #Input Products
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[2]/div[1]').click() #Search
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr/td[9]/span[1]/div/img').click() #Actions Button
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[4]/div/div[1]/div/div[2]/table/tbody/tr/td/div[2]/ul/li/div/span/span[1]/div[1]/img').click()  # PACK Button
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div/div[2]/div[4]/span').click()  # Add Serial Button
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[1]/div/div/div/div[5]/div/div/div[1]/div/div[1]/table/thead/tr/th[1]/input').click()  # Select All Serial
        self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[2]/div/button[1]/span').click()  # Add Button
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button/span').click()  #Close Button
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[2]/span').click()  #Save Button
        sleep(2)

        #Task - Structure that Checks the "Done" Status of the Task
        self.navegador.get(self.Link + "/utilities/background_tasks")
        sleep(2)
        while len(self.navegador.find_elements(By.XPATH, "/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr[1]/td[5]//span[text()='Done']")) != 1:
            self.navegador.refresh()
            sleep(2)

        # START OF THE "UNPACKING" SCREEN
        self.navegador.get(self.Link + "/unpacking/")
        sleep(2)
        self.navegador.get(self.Link + "/unpacking/search")
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[1]/form/div[1]/div[10]/div[2]/input').send_keys("FATHERLOT") #Input Products
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[2]/div[1]').click() #Search
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr/td[10]/span[1]/div/img').click() #Select Unpack
        sleep(2)

        # GENERATES SECOND PART OF THE NDC, COMPOSED OF 03 DIGITS.
        ph_no = []
        # A loop is used to generate 5 Numbers.
        for i in range(0, 4):
            ph_no.append(r.randint(0, 9))  # Each Number from 0 to 9
        # I take the numbers from the array and put them together
        REF = ''
        for i in ph_no:
            REF += str(i)

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/input').send_keys("REF#" + REF) #Input Products
        Select(self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[3]/div[2]/div/select')).select_by_index(3) #Select Reason
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[7]/div/div[1]/div/div[2]/table/tbody/tr/td/div[2]/ul/li/div/i').click() #Open Products
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click() #Confirm Storage Products
        sleep(2)
        Select(self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div/div/div/div[1]/div/div[2]/table/tbody/tr/td[3]/span[1]/select')).select_by_index(1) #Select Storage Area
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click() #Unpack Button
        sleep(2)

        #START OF THE DISPORSAL
        self.navegador.get(self.Link + "/adjustments/destructions")  #Link
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[2]/div/div/span').click()  # Button Destruct
        sleep(2)

        #Change Location
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[1]/div[3]/a').click()  # Change Location
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[1]/form/div[1]/div[2]/div[2]/input').send_keys(self.Location)  # Search Location
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[2]/div').click()  # Button Search
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[2]/table/tbody/tr/td[3]/div/img').click()  # Button Search
        sleep(2)

        Select(self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div[2]/div[1]/div[2]/div/select')).select_by_index(1)  # Reason
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[2]/span').click()  #Items Tab
        sleep(2)

        #Child Lot
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[2]/div/div[2]/div[4]/span').click()  #Add Look Up
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[1]/form/div[1]/div[3]/div[2]/input').send_keys(NDCOFICIAL_CHILD)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[2]/div[1]').click()  #Search
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[2]/table/tbody/tr[1]/td[6]/div[2]/img').click()  #Product
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[4]/div/div/div[1]/div/div[2]/table/tbody/tr[1]/td[4]/span[1]/div/img').click()  #Lot 01
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[1]/div/div/div/div[5]/div/div/div[1]/div/div[1]/table/thead/tr/th[1]/input').click()  #All Serials
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[2]/div/button[1]/span').click()  #Add Selection
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click()  #Ok in New Destructions
        sleep(2)

        #Task - Structure that Checks the "Done" Status of the Task
        self.navegador.get(self.Link + "/utilities/background_tasks")
        sleep(2)
        while len(self.navegador.find_elements(By.XPATH, "/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr[1]/td[5]//span[text()='Done']")) != 1:
            self.navegador.refresh()
            sleep(2)

        #START IN THE DELETE PRODUCT
        self.navegador.get(self.Link + "/products/")  #Link
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[1]/form/div[1]/div[1]/div[2]/input').send_keys(NDCOFICIAL_FATHER)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[2]/div[1]').click()  #Search
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr/td[7]/span[1]/div[3]/img').click()  #Button Delete
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click()  # Confirm Delete
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[1]/form/div[1]/div[1]/div[2]/input').clear()
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[1]/form/div[1]/div[1]/div[2]/input').send_keys(NDCOFICIAL_CHILD)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[2]/div[1]').click()  #Search
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr/td[7]/span[1]/div[3]/img').click()  # Button Delete
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click()  # Confirm Delete
        sleep(2)

        #START IN THE DELETE SERIALIZED LOT
        self.navegador.get(self.Link + "/products/manufacturer")  #Link
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[1]/form/div[1]/div[3]/div[2]/input').send_keys("CHILDLOT")
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[2]/div[1]').click()  #Search
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr/td[9]/span[1]/div[3]/img').click()  #Button Delete
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click()  # Confirm Delete
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[1]/form/div[1]/div[3]/div[2]/input').clear()
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[1]/form/div[1]/div[3]/div[2]/input').send_keys("FATHERLOT")
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[2]/div[1]').click()  #Search
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr/td[9]/span[1]/div[3]/img').click()  #Button Delete
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click()  # Confirm Delete

        self.End_Time_Module = datetime.now().strftime("%I:%M %p")
        self.Result = "Successfully"
        print("Commission and Packaging Complete!")

    def company_management(self):

        self.Module = "Company Management"
        self.Start_Time_Module = datetime.now().strftime("%I:%M %p")

        #START OF THE COMPANY MANAGEMENT

        self.navegador.get(self.Link + "/company_mgt/")  #Link
        sleep(2)

        #START OF THE COMPANY SETTINGS SCREEN

        #General Tab
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div[1]/ul/li[1]/a').click()  #Button
        sleep(4)
        Select(self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div[2]/div[1]/div/select')).select_by_index(1) #Timezone
        Select(self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[3]/div[2]/div[1]/div/select')).select_by_index(1) #Date Format
        Select(self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[4]/div[2]/div[1]/div/select')).select_by_index(1) #Time Format
        Select(self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[5]/div[2]/div[1]/div/select')).select_by_index(1) #Language
        Select(self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[6]/div[2]/div[1]/div/select')).select_by_index(1) #Dashboard

        #Security Tab
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[2]/span').click()  #Change the Tab
        sleep(2)

        #Policies Tab
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[3]/span').click()  #Change the Tab
        sleep(2)

        #Products Tab
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[4]/span').click()  #Change the Tab
        sleep(2)

        #Inventory Tab
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[5]/span').click()  #Change the Tab
        sleep(2)

        #Misc Tab
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[6]/span').click()  #Change the Tab
        sleep(2)

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[2]/span').click()  #Button Cancel
        sleep(2)

        #RETURN TO THE COMPANY MENU

        #START OF THE LOCATIONS SCREEN
        self.navegador.get(self.Link + "/company_mgt/locations") #Link
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[2]/div/div[3]/span').click() #Add Button
        sleep(2)

        #Generating Location Suffix
        ph_no = []
        #A loop is used to generate 5 Numbers.
        for i in range(0, 5):
            ph_no.append(r.randint(0, 9))  #Each Number from 0 to 9
        #I take the numbers from the array and put them together
        LC = ''
        for i in ph_no:
            LC += str(i)

        #General Tab
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[1]/div[2]/input').send_keys("Location " + LC) #Location Name
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div[2]/input').send_keys("Location " + LC + " - Location Detail") #Location Name

        #Generating the First Part of the GLN
        ph_no = []
        #A loop is used to generate 7 Numbers.
        for i in range(0, 7):
            ph_no.append(r.randint(0, 9))  #Each Number from 0 to 9
        #I take the numbers from the array and put them together
        GLN1 = ''
        for i in ph_no:
            GLN1 += str(i)

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[3]/div[2]/input').send_keys(GLN1 + "000001") #GLN
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[4]/div[2]/input').send_keys("urn:epc:id:sgln:" + GLN1 + ".000000.0") #SGLN

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click() #Add Form Button
        sleep(2)

        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[1]/form/div[1]/div[2]/div[2]/input').send_keys("Location " + LC) #Location Search Field
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[2]/div[1]').click() #Search Button
        sleep(4)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr[1]/td[8]/span[1]/div[2]/img').click() #Edit Button

        #Address Tab
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[2]/span').click() #Change the Tab
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div/div[2]/div/div[1]/span').click() #Add Button
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[3]/div[2]/input').send_keys("Location " + LC + " Address") #Nickname
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[4]/div[2]/input').send_keys(GLN1 + "000001") #GLN
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[5]/div[2]/input').send_keys("urn:epc:id:sgln:" + GLN1 + "000000.0") #SGLN
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[6]/div[2]/input').send_keys("Test") #Recipient Name
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[7]/div[2]/input').send_keys("400 Av Washington") #Line 1
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[13]/div[2]/input').send_keys("Montgomery") #City
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[14]/div[2]/input').send_keys("36104") #Zip Code
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click() #Add Button
        sleep(2)

        #RETURN TO THE COMPANY MENU
        self.navegador.get(self.Link + "/company_mgt") #Link
        sleep(2)

        #START OF THE THIRD PARTY LOGISTICS
        self.navegador.get(self.Link + "/company_mgt/third_party_logistics_providers") #Link
        sleep(4)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[2]/div/div/span').click() #Add Button
        sleep(2)

        #General Tab
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[1]/div[2]/input').send_keys("Test") #Name 3PL
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[6]/div[2]/div/div/div[2]/div/div/span').click() #Add Button
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[1]/div[1]/div[2]/input').send_keys("Test") #Name 3PL Location - General Tab
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/ul/li[2]/span').click() #Change Address Tab
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[2]/div[1]/div[2]/input').send_keys("Test") #Recipient Name - Adress Tab
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[2]/div[2]/div[2]/input').send_keys("Test") #Adress - Adress Tab
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[2]/div[6]/div[2]/input').send_keys("Test") #City - Adress Tab
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[2]/div[9]/div[2]/input').send_keys("Test") #Zip Code - Adress Tab
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click() #Add Button
        sleep(2)

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click() #Add Button
        sleep(2)

        #RETURN TO THE COMPANY MENU
        self.navegador.get(self.Link + "/company_mgt") #Link
        sleep(2)

        #START OF THE WORKFLOW AUTOMATION
        self.navegador.get(self.Link + "/company_mgt/workflow_automation") #Link
        sleep(4)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[2]/div/div/span').click() #Add Button
        sleep(2)

        #General Tab
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[1]/div[2]/input').send_keys("Test") #Name 3WA

        #Conditions Tab
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[2]/span').click() #Change Tab
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div/div[2]/div/div/span').click() #Add Conditions
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click() #Add in Add Conditions
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/input').send_keys("Test") #Name file Condition
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[2]/div/button[1]/span').click() #Ok in file Condition
        sleep(2)

        #Triggers Tab
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[3]/span').click() #Change Tab
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[3]/div/div[2]/div/div/span').click() #Add Triggers
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click() #Add in Add Trigger
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/input').send_keys("Test") #Name Inbound Triggers
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[2]/div/button[1]/span').click() #Ok in Inbound Triggers
        sleep(4)

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click() #Add Button
        sleep(2)

        #RETURN TO THE COMPANY MENU
        self.navegador.get(self.Link + "/company_mgt") #Link
        sleep(2)

        #START OF THE STAFF MANAGEMENT
        self.navegador.get(self.Link + "/company_mgt/staff_management") #Link
        sleep(4)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[2]/div/div/span').click() #Add Button
        sleep(2)

        #Generating Location Suffix
        ph_no = []
        #A loop is used to generate 5 Numbers.
        for i in range(0, 5):
            ph_no.append(r.randint(0, 9))  #Each Number from 0 to 9
        #I take the numbers from the array and put them together
        UsStaff = ''
        for i in ph_no:
            UsStaff += str(i)

        #General Tab
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[1]/div[2]/input').send_keys("This is a Test User") #Name 3WA
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div[2]/input').send_keys("Administrator") #Name 3WA
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[3]/div[2]/input').send_keys("test" + UsStaff + "@test.com") #Name 3WA
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[5]/div/div/div[3]/div[2]/input').send_keys("TrackTraceRX2022") #Name 3WA
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[5]/div/div/div[4]/div[2]/div/input').click() #Force Password Change
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click() #Add in Create User
        sleep(2)

        #Users Screen
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[1]/form/div[1]/div[2]/div[2]/input').send_keys("test" + UsStaff + "@test.com") #Search Username Field
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[2]/div[1]').click() #Search Button
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr/td[7]/span[1]/div[2]/img').click() #Edit User
        sleep(2)

        #Roles Tab
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[2]/span').click() #Change Tab
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div/div[2]/div/div/span').click() #Add Role
        sleep(5)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click() #Add in Add Role
        sleep(2)

        #Permissions Tab
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[3]/span').click() #Change Tab
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[3]/div/div[2]/div/div/span').click() #Add Permission
        sleep(5)
        Select(self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[2]/div[2]/div/select')).select_by_index(2) #Access Level in add Permission
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click() #Add in Add Permission
        sleep(2)

        #RETURN TO THE COMPANY MENU.
        self.navegador.get(self.Link + "/company_mgt") #Link
        sleep(4)

        #START OF THE API ACCESS KEYS.
        self.navegador.get(self.Link + "/company_mgt/api_keys") #Link
        sleep(4)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[2]/div/div/span').click() #Add Button
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/input').send_keys("Test") #Description
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click() #Add Button
        sleep(2)

        #IMMUTABLE FOOTER. MAKE CHANGES ONLY UP TO HERE.

        #RETURN TO THE COMPANY MENU
        self.navegador.get(self.Link + "/company_mgt") #Link

        self.End_Time_Module = datetime.now().strftime("%I:%M")
        self.Result = "MODULE EXECUTED SUCCESSFULLY"
        print("Company Mgt Complete!")

    def disporsal(self):

        self.Module = "Disporsal"
        self.Start_Time_Module = datetime.now().strftime("%I:%M %p")

        #START OF THE DISPORSAL
        self.navegador.get(self.Link + "/adjustments/destructions")  #Link
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[2]/div/div/span').click()  # Button Destruct
        sleep(2)

        #Change Location
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[1]/div[3]/a').click()  # Change Location
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[1]/form/div[1]/div[2]/div[2]/input').send_keys(self.Location)  # Search Location
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[2]/div').click()  # Button Search
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[2]/table/tbody/tr/td[3]/div/img').click()  # Select Location

        Select(self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div[2]/div[1]/div[2]/div/select')).select_by_index(1)  # Reason
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[2]/span').click()  #Items Tab
        sleep(2)

        Tb = pd.read_excel(self.file)
        for i, product in enumerate(Tb["product_name"]):
            NDCOFICIAL = Tb.loc[i, "ndc"]
            LOT_NAME = Tb.loc[i, "lot_name"]

            self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[2]/div/div[2]/div[4]/span').click()  # Add Look Up
            sleep(2)
            self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[1]/form/div[1]/div[3]/div[2]/input').send_keys(NDCOFICIAL)
            self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[2]/div[1]').click()  # Pesquisar
            sleep(2)
            self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[2]/table/tbody/tr/td[6]/div[2]/img').click()  # Select Product
            sleep(2)
            self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[4]/div/div/div[1]/div/div[2]/table/tbody/tr[1]/td[4]/span[1]/div/img').click()  # Select Product
            sleep(2)
            self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[1]/div/div/div/div[5]/div/div/div[1]/div/div[1]/table/thead/tr/th[1]/input').click()  # Select All Serials
            self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[2]/div/button[1]/span').click()  # Click Add Selection
            sleep(2)


        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click()  #Ok in New Destructions
        sleep(2)

        #Task - Structure that Checks the "Done" Status of the Task
        self.navegador.get(self.Link + "/utilities/background_tasks")
        sleep(2)
        while len(self.navegador.find_elements(By.XPATH, "/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr[1]/td[5]//span[text()='Done']")) != 1:
            self.navegador.refresh()
            sleep(2)

        self.End_Time_Module = datetime.now().strftime("%I:%M")
        self.Result = "Successfully"
        print("Destruction Complete!")

    def transformation(self):

        self.Module = "Transformation"
        self.Start_Time_Module = datetime.now().strftime("%I:%M %p")

        #START OF THE "MANAGEMENT PRODUCTS" SCREEN. LET'S ADD 02 PRODUCTS, AN INGREDIENT AND AN OUTCOME.

        #CREATE PRODUCT NINGREDIENT

        self.navegador.get(self.Link + "/products/")  #Product Management
        sleep(4)

        # GENERATES THE PREFIX, THE TWO NUMBERS THAT PRECED THE FIRST PART OF THE NDC IN THE COMPANY PREFIX.
        ph_no = []
        # A loop is used to generate 2 Numbers.
        for i in range(0, 2):
            ph_no.append(r.randint(0, 9))  # Each Number from 0 to 9
        # I take the numbers from the array and put them together
        W1 = ''
        for i in ph_no:
            W1 += str(i)

        # GENERATES FIRST PART OF THE NDC, COMPOSED OF 05 DIGITS.
        ph_no = []
        # A loop is used to generate 5 Numbers.
        for i in range(0, 5):
            ph_no.append(r.randint(0, 9))  # Each Number from 0 to 9
        # I take the numbers from the array and put them together
        X1 = ''
        for i in ph_no:
            X1 += str(i)

        # GENERATES SECOND PART OF THE NDC, COMPOSED OF 03 DIGITS.
        ph_no = []
        # A loop is used to generate 3 Numbers.
        for i in range(0, 3):
            ph_no.append(r.randint(0, 9))  # Each Number from 0 to 9
        # I take the numbers from the array and put them together
        Y1 = ''
        for i in ph_no:
            Y1 += str(i)

        # GENERATES SECOND PART OF THE NDC, COMPOSED OF 03 DIGITS.
        ph_no = []
        # A loop is used to generate 5 Numbers.
        for i in range(0, 2):
            ph_no.append(r.randint(0, 9))  # Each Number from 0 to 9
        # I take the numbers from the array and put them together
        Z1 = ''
        for i in ph_no:
            Z1 += str(i)

        # GENERATES PRODUCT NUMBER
        W = W1
        X = X1
        Y = Y1
        Z = Z1
        GS1PREFIX = W + X
        GS1ID = "0" + Y + Z
        NDCOFICIAL_I = X + "-" + Y + "-" + Z

        #Add Products Test
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[2]/div/div[3]/span').click()  #Add
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div[3]/div[2]/input').send_keys("Product_Ingredient")  # Campo Nome Produto
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[2]/span').click()  #Identifier Tab
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[1]/div[2]/input').send_keys("9999999999")  #SKU
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[2]/div[2]/input').send_keys("888888888888")  #UPC
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[3]/div[2]/input').send_keys(GS1PREFIX)  #GS1 Company Prefix
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[4]/div[2]/input').send_keys(GS1ID)  #GS1 ID
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[5]/div[2]/div[2]/div/div/span').click()  #Add Identifiers
        sleep(2)
        Select(self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div/select')).select_by_index(0)  #Select the First ComboBox Option
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[2]/div[2]/input').send_keys(NDCOFICIAL_I)  #NDC
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click()  #Add Identifier
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click()  #Add button to Product Form

        #CREATE PRODUCT OUTCOME

        # GENERATES THE PREFIX, THE TWO NUMBERS THAT PRECED THE FIRST PART OF THE NDC IN THE COMPANY PREFIX.
        ph_no = []
        # A loop is used to generate 2 Numbers.
        for i in range(0, 2):
            ph_no.append(r.randint(0, 9))  # Each Number from 0 to 9
        # I take the numbers from the array and put them together
        W1 = ''
        for i in ph_no:
            W1 += str(i)

        # GENERATES FIRST PART OF THE NDC, COMPOSED OF 05 DIGITS.
        ph_no = []
        # A loop is used to generate 5 Numbers.
        for i in range(0, 5):
            ph_no.append(r.randint(0, 9))  # Each Number from 0 to 9
        # I take the numbers from the array and put them together
        X1 = ''
        for i in ph_no:
            X1 += str(i)

        # GENERATES SECOND PART OF THE NDC, COMPOSED OF 03 DIGITS.
        ph_no = []
        # A loop is used to generate 3 Numbers.
        for i in range(0, 3):
            ph_no.append(r.randint(0, 9))  # Each Number from 0 to 9
        # I take the numbers from the array and put them together
        Y1 = ''
        for i in ph_no:
            Y1 += str(i)

        # GENERATES SECOND PART OF THE NDC, COMPOSED OF 03 DIGITS.
        ph_no = []
        # A loop is used to generate 5 Numbers.
        for i in range(0, 2):
            ph_no.append(r.randint(0, 9))  # Each Number from 0 to 9
        # I take the numbers from the array and put them together
        Z1 = ''
        for i in ph_no:
            Z1 += str(i)

        # GENERATES PRODUCT NUMBER
        W = W1
        X = X1
        Y = Y1
        Z = Z1
        GS1PREFIX = W + X
        GS1ID = "0" + Y + Z
        NDCOFICIAL_O = X + "-" + Y + "-" + Z

        #Add Products Test
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[2]/div/div[3]/span').click()  #Add
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div[3]/div[2]/input').send_keys("Product_Outcome")  # Campo Nome Produto
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[2]/span').click()  #Identifier Tab
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[1]/div[2]/input').send_keys("9999999999")  #SKU
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[2]/div[2]/input').send_keys("888888888888")  #UPC
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[3]/div[2]/input').send_keys(GS1PREFIX)  #GS1 Company Prefix
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[4]/div[2]/input').send_keys(GS1ID)  #GS1 ID
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[5]/div[2]/div[2]/div/div/span').click()  #Add Identifiers
        sleep(2)
        Select(self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div/select')).select_by_index(0)  #Select the First ComboBox Option
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[2]/div[2]/input').send_keys(NDCOFICIAL_O)  #NDC
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click()  #Add Identifier
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click()  #Add button to Product Form

        #START RECEIVING

        #Receiving
        self.navegador.get(self.Link + "/receiving") #Receiving Link
        sleep(2)

        #Inbound
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[2]/div/div[2]/span').click() #Manual Inbound
        sleep(2)

        #PO Menu
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div[2]/div[3]/a').click() #Change Seller
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[1]/form/div[1]/div[2]/div[2]/input').send_keys(self.Trading_Proc) #Search TP
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[2]/div').click() #Button Search
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[2]/table/tbody/tr[1]/td[3]/div[2]/img').click() #Select TP

        # Generating PO and SO Identifications
        ph_no = []
        # A loop is used to generate 5 Numbers.
        for i in range(0, 5):
            ph_no.append(r.randint(0, 9))  # Each Number from 0 to 9
            # I take the numbers from the array and put them together
        Purchase_01 = ''
        for i in ph_no:
            Purchase_01 += str(i)
        self.PO_Name_Serial = "PO_Teste_" + Purchase_01

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div[4]/div[2]/input').send_keys(self.PO_Name_Serial) #Customer Order ID
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div[5]/div[2]/input').send_keys(self.PO_Name_Serial) #Invoice
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div[6]/div[2]/input').send_keys(self.PO_Name_Serial) #PO Number
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div[4]/div[4]/input').send_keys(self.PO_Name_Serial) #Internal Reference
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div[5]/div[4]/input').send_keys(self.PO_Name_Serial) #Release
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div[6]/div[4]/input').send_keys(self.PO_Name_Serial) #Order Number

        #Sold By and Ship From Tab
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[2]/span').click()  #Sold By

        #Bought By and Ship To Tab
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[3]/span').click()  #Bought By

        # # Shipment Tab
        # self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[4]/span').click()  # Shipment Tab
        # sleep(2)
        # Select(self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[4]/div[6]/div[2]/div[1]/div[2]/div[3]/div/select')).select_by_index(1)  # Select Lot Select

        #Itens Tab
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[5]').click() #Line Items
        sleep(1)

        #Enter Serial Based Screen
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[5]/div/div[1]/div[2]/div/div/span').click() #Add Product
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[1]/form/div[1]/div[3]/div[2]/input').send_keys(NDCOFICIAL_I) #NDC Search
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[2]/div[1]').click() #Button to search for informed NDC
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[2]/table/tbody/tr/td[6]/div[2]/img').click() #Click in Product
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[3]/div[2]/input').send_keys(1) #Quantity
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click() #Ok in Quantity Form
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[3]/div/div/div[3]/div[2]/div/div/span').click() #Add Lot
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[1]/div/div/div/div[1]/div[3]/div[2]/input').send_keys("LOT_TRANSFORMATION") #Lot Name
        self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[1]/div/div/div/div[1]/div[4]/div[2]/div[1]/div/input').click() #Radio Serial Based
        self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[1]/div/div/div/div[1]/div[3]/div[4]/input').send_keys("08-08-2024") #Expiration
        self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[2]/div/button[1]/span').click() #Ok in Add Lot
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click() #Ok in Product Information

        #Aggregation Menu

        #Aggregation Serials
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[6]').click() #Aggregation Tab
        sleep(2)

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[6]/div[2]/div/div[2]/div/div/span').click()  # Add Aggregation
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[4]/div/div/input').click()  # Select Product Aggregation
        Select(self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[5]/div[3]/div/select')).select_by_index(1)  # Select Product Select
        Select(self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[6]/div[3]/div/select')).select_by_index(1)  # Select Lot Select

        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[8]/div[2]/div/div[3]/div[2]/textarea').send_keys("Serial_T_001" + "\n")  # Type inside the Serials TextArea.

        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click()  # OK in Aggregation

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click()  #OK in Form Inbound
        sleep(5)

        #Task - Structure that Checks the "Done" Status of the Task
        self.navegador.get(self.Link + "/utilities/background_tasks")
        sleep(2)
        while len(self.navegador.find_elements(By.XPATH, "/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr[1]/td[5]//span[text()='Done']")) != 1:
            self.navegador.refresh()
            sleep(2)

        #START REPACKER

        #Recipes Management
        self.navegador.get(self.Link + "/transformation/recipes") #Receiving Link
        sleep(2)

        ph_no = []
        # A loop is used to generate 5 Numbers.
        for i in range(0, 5):
            ph_no.append(r.randint(0, 9))  # Each Number from 0 to 9
            # I take the numbers from the array and put them together
        R = ''
        for i in ph_no:
            R += str(i)

        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[2]/div/div/span').click()  # Add Recipe
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[1]/div[2]/input').send_keys("Recipe_" + str(R))  # Add Recipe
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[3]/span').click()  # Ingredients Tab
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[3]/div[3]/div[2]/div/div/span').click()  # Add Product
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click()  # Ok in Add Input Product
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[1]/form/div[1]/div[3]/div[2]/input').send_keys(NDCOFICIAL_I)  # Input NDC
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[2]/div[1]').click()  # Search Button
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[2]/table/tbody/tr/td[6]/div[2]/img').click()  # Select Product
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[2]/div[2]/input').send_keys(1)  # Input Quantity
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click()  # Add Specific Input Product
        sleep(2)

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[4]/span').click()  # Outcome Products
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[4]/div[4]/div[2]/div/div/span').click()  # Add Product
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[1]/form/div[1]/div[3]/div[2]/input').send_keys(NDCOFICIAL_O)  # Input NDC
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[2]/div[1]').click()  # Search Button
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div[1]/div[2]/div/div/div/div/div/div[2]/table/tbody/tr/td[6]/div[2]/img').click()  # Select Product
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[2]/div[3]/div[2]/input').send_keys(3)  # Input NDC
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click()  # Add Specific Input Product
        sleep(2)

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click()  # Add in Recipe
        sleep(2)

        #Transform Inventory

        self.navegador.get(self.Link + "/transformation/") #Receiving Link
        sleep(2)

        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/ul/li[1]/a').click()  # Transform Inventory
        sleep(2)

        Select(self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div/select')).select_by_visible_text("Recipe_" + str(R))  # Select Product Select
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click()  # Ok in Transform Product
        sleep(2)

        #Transform Log

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[2]/span').click()  # Step Tab
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[5]/div[2]/div/input').click()  # Choose "Transforming" Step

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[3]/span').click()  # Ingredients Tab
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[3]/div/div[1]/div/div[2]/table/tbody/tr/td[6]/span[1]/div[2]/img').click()  # Ingredients Records
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[3]/div[2]/div/div[1]/span').click()  # Add Ingredient
        sleep(4)
        self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[1]/div/div/div/div[3]/div/div/div[1]/div/div[2]/table/tbody/tr[1]/td[4]/span[1]/div/img').click()  # Select Lot
        sleep(4)
        self.navegador.find_element(By.XPATH, '/html/div[8]/div/div[2]/div[1]/div/div/div/div[4]/div/div/div[1]/div/div[2]/table/tbody/tr[1]/td[1]/input').click()  # Select Serial
        self.navegador.find_element(By.XPATH, '/html/div[8]/div/div[2]/div[2]/div/button[1]/span').click()  # Add Selection
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click()  # Ok in Transform Product
        sleep(2)

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[4]/span').click()  # Outcome Tab
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[4]/div/div[1]/div/div[2]/table/tbody/tr/td[6]/span[1]/div[3]/img').click()  # Ingredient Record
        sleep(2)
        Select(self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[1]/div[8]/div[2]/div/select')).select_by_index(1)  # Select Product Select
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click()  # Ok in Outcome Products
        sleep(2)

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click()  # Ok in Transform Product
        sleep(2)

        #Task - Structure that Checks the "Done" Status of the Task
        self.navegador.get(self.Link + "/utilities/background_tasks")
        sleep(2)
        while len(self.navegador.find_elements(By.XPATH, "/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr[1]/td[5]//span[text()='Done']")) != 1:
            self.navegador.refresh()
            sleep(2)

        #Recipes Log
        self.navegador.get(self.Link + "/transformation/inventory_transformation_log")  # Inventory Link
        sleep(2)

        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[2]/div[2]').click()  # Advanced Search
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[1]/form/div[2]/div[2]/div[2]/input').send_keys("Recipe_" + str(R))  # Add Selection
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[2]/div[1]').click()  # Search Button
        sleep(2)

        #Edit Step
        self.navegador.find_element(By.XPATH, '//html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr/td[8]/span[1]/div[2]/img').click()  # Edit Button
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[2]/span').click()  # Step Tab
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[7]/div[2]/div/input').click()  # Select Quality control

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/ul/li[4]/span').click()  # Outcome Tab
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[4]/div/div[1]/div/div[2]/table/tbody/tr/td[6]/span[1]/div[1]/img').click()  # View Serials
        sleep(2)

        #Pega os Seriais Gerados
        elemento_span = self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[2]/div/div[1]/div/div[2]/table/tbody/tr[1]/td[3]/span[1]')
        Serial_01 = elemento_span.text
        elemento_span = self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[2]/div/div[1]/div/div[2]/table/tbody/tr[2]/td[3]/span[1]')
        Serial_02 = elemento_span.text
        elemento_span = self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[2]/div/div[1]/div/div[2]/table/tbody/tr[3]/td[3]/span[1]')
        Serial_03 = elemento_span.text
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button/span').click()  # Cancel Button
        sleep(2)

        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[1]/div/div/div/div[4]/div/div[1]/div/div[2]/table/tbody/tr/td[6]/span[1]/div[3]/img').click()  # Ingredients Record
        sleep(2)

        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/ul/li[4]/span').click()  # Sampling Tab
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[4]/section/div[1]/div[4]/a').click()  # Select Serials
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[1]/div/div/div/div[2]/div/textarea').send_keys(Serial_01) #Outcome Sample
        self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[2]/div/button[1]/span').click()  # Ok in Outcome Sample
        sleep(2)
        Select(self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[4]/section/div[3]/div[3]/div/select')).select_by_index(1)  # Select Area

        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/ul/li[5]/span').click()  # Quality Control
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[5]/div[3]/div/div/input').click()  # Accepted

        formato_personalizado = "%m/%d/%Y"
        data_atual = datetime.now()
        data_formatada = data_atual.strftime(formato_personalizado)

        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[5]/section/div/div[2]/div[2]/div[1]/input').send_keys(data_formatada)  # Date Now

        data_formatada = datetime.now().strftime("%I:%M")

        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[5]/section/div/div[2]/div[2]/div[3]/input').send_keys(data_formatada)  # Hours Now

        data_formatada = datetime.now().strftime("%p")

        Select(self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[5]/section/div/div[2]/div[2]/div[4]/div/select')).select_by_value(data_formatada)  # AM or PM

        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[5]/section/div/div[3]/div[2]/div[2]/a').click()  # Select Serial Consumed
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[1]/div/div/div/div[2]/div/textarea').send_keys(Serial_02)
        self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[2]/div/button[1]/span').click()

        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[5]/section/div/div[4]/div[2]/div[2]/a').click()  # Select Serial Remove
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[1]/div/div/div/div[2]/div/textarea').send_keys(Serial_03)
        self.navegador.find_element(By.XPATH, '/html/div[6]/div/div[2]/div[2]/div/button[1]/span').click()

        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[5]/section/div/div[5]/div[2]/input').send_keys("QA Test")
        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[1]/div/div/div/div[5]/section/div/div[6]/div[2]/input').send_keys("QA Test")

        self.navegador.find_element(By.XPATH, '/html/div[4]/div/div[2]/div[2]/div/button[1]/span').click() #Ok in Outcome Products
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click() #Ok in Transform Product

        #Task - Structure that Checks the "Done" Status of the Task
        self.navegador.get(self.Link + "/utilities/background_tasks")
        sleep(2)
        while len(self.navegador.find_elements(By.XPATH, "/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr[1]/td[5]//span[text()='Done']")) != 1:
            self.navegador.refresh()
            sleep(2)

        # START IN THE DELETE PRODUCT
        self.navegador.get(self.Link + "/products/")  # Link
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[1]/form/div[1]/div[1]/div[2]/input').send_keys(NDCOFICIAL_I)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[2]/div[1]').click()  # Search
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr/td[7]/span[1]/div[3]/img').click()  # Button Delete
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click()  # Confirm Delete
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[1]/form/div[1]/div[1]/div[2]/input').clear()
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[1]/form/div[1]/div[1]/div[2]/input').send_keys(NDCOFICIAL_O)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[1]/div[2]/div[1]').click()  # Search
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/body/div[1]/div[5]/div[2]/div/div/div[1]/div[3]/table/tbody/tr/td[7]/span[1]/div[3]/img').click()  # Button Delete
        sleep(2)
        self.navegador.find_element(By.XPATH, '/html/div[2]/div/div[2]/div[2]/div/button[1]/span').click()  # Confirm Delete
        sleep(2)

        self.End_Time_Module = datetime.now().strftime("%I:%M %p")
        self.Result = "Successfully"
        print("Transformation Complete")

    def dropship(self):

        print("A")

    # Functions to build the PDF Report
    def pdf(self):

        def add_text(output_file):
            Portal = "QA Assurance"

            ph_no = []
            # A loop is used to generate 5 Numbers.
            for i in range(0, 5):
                ph_no.append(r.randint(0, 9))  # Each Number from 0 to 9
            # I take the numbers from the array and put them together
            R = ''
            for i in ph_no:
                R += str(i)

            ph_no = []
            # A loop is used to generate 5 Numbers.
            for i in range(0, 5):
                ph_no.append(r.randint(0, 9))  # Each Number from 0 to 9
            # I take the numbers from the array and put them together
            self.P = ''
            for i in ph_no:
                self.P += str(i)

            Reference_Report = "ReportRPA_" + R
            Date = datetime.now().strftime("%m/%d/%Y")
            Author = "Victor Angelo"

            Path = getattr(sys, '_MEIPASS', os.path.abspath(os.path.dirname(__file__)))
            output_path = Path + "\\Archives\\Detailed\\Track_Validation_" + self.P + ".pdf"

            # Create 01 Page Document
            cnv = canvas.Canvas(output_path, pagesize=A4)
            Path = getattr(sys, '_MEIPASS', os.path.abspath(os.path.dirname(__file__)))
            pdfmetrics.registerFont(
                TTFont('Arial', 'arial.ttf'))  # Replace 'arial.ttf' with the actual path to your Arial font file

            # Add Logo
            image_path = os.path.join(Path + "\\Archives\\images\\", "Logo_TrackTraceRX.png")
            cnv.drawImage(image_path, 300, 690, width=250, height=40)

            # Text = TrackRX Validation Report for
            font_name = "Arial"
            font_size = 24
            cnv.setFont(font_name, font_size)
            cnv.drawString(50, 550, "TrackRX Validation Report for")
            cnv.drawString(50, 525, Portal)

            # Text = Reference Report
            font_size = 13.5
            cnv.setFont(font_name, font_size)
            cnv.drawString(50, 350, "Report Reference Number:")
            cnv.drawString(220, 350, Reference_Report)

            # Text = Date Report
            cnv.drawString(50, 330, "Date of Issue:")
            cnv.drawString(140, 330, Date)

            # Text = Author RPA
            cnv.drawString(50, 310, "Author:")
            cnv.drawString(100, 310, Author)

            # Save 01 Page Document
            cnv.showPage()

            # Text = Execultive Summary
            font_name = "Arial"
            font_size = 16
            cnv.setFont(font_name, font_size)
            cnv.drawString(50, 750, "1. Execultive Summary")

            # Text = Text of Execultive Summary
            font_size = 12
            cnv.setFont(font_name, font_size)
            cnv.drawString(50, 720,
                           "This report details the validation activities performed from " + Date + " for the Customers")
            cnv.drawString(50, 700, "Portal capturing findings from the RPA phases.")
            cnv.drawString(50, 680, "The TrackRX Portal successfully met all validation criteria.")

            # Text = Introduction Summary
            font_name = "Arial"
            font_size = 16
            cnv.setFont(font_name, font_size)
            cnv.drawString(50, 620, "2. Introduction")

            # Text = Text of Introduction Summary
            font_size = 12
            cnv.setFont(font_name, font_size)
            cnv.drawString(50, 590,
                           "The validation activities are intended to confirm that the TrackRX Portal is installed, operates,")
            cnv.drawString(50, 570, "and performs according to manufacturer specifications and company requirements.")

            # Text = Deviations and Corrective Actions Summary
            font_name = "Arial"
            font_size = 16
            cnv.setFont(font_name, font_size)
            cnv.drawString(50, 510, "3. Deviations and Corrective Actions")

            # Text = Text of Deviations and Corrective Actions Summary
            font_size = 12
            cnv.setFont(font_name, font_size)
            cnv.drawString(50, 480, "No deviations were observed during the validation process.")

            # Text = Methodology Summary
            font_name = "Arial"
            font_size = 16
            cnv.setFont(font_name, font_size)
            cnv.drawString(50, 420, "4. Methodology")

            # Text = Text of Deviations and Corrective Actions Summary
            font_size = 12
            cnv.setFont(font_name, font_size)
            cnv.drawString(50, 390, "You can find the entire detailed process that was performed by Track Trace RPA on the")
            cnv.drawString(50, 370, "next page.")

            # Save 02 Page Document
            cnv.showPage()

            # Text = Process Sub Summary
            font_name = "Arial"
            font_size = 16
            cnv.setFont(font_name, font_size)
            cnv.drawString(50, 750, "5. Detailed RPA Process")

            font_name = "Arial"
            font_size = 12
            cnv.setFont(font_name, font_size)

            Path = getattr(sys, '_MEIPASS', os.path.abspath(os.path.dirname(__file__)))
            input_txt = Path + "\\Archives\\Output\\" + self.f
            # Adicione o conteúdo do arquivo de texto ao novo PDF
            with open(input_txt, 'r', encoding='latin-1') as txt_file:
                lines = txt_file.readlines()

                espacamento_entre_linhas = 20
                y_position = 720
                for line in lines:
                    cnv.drawString(50, y_position, line.strip())
                    y_position -= espacamento_entre_linhas  # Adjust to next line spacing

            # Save 03 Page Document and Save Document
            cnv.save()

        add_text("Track_Validation.pdf")

class App(Functions):

    #Method Initialized When Calling the Class App

    def __init__(self):
        self.TWBD()
        self.Location = "Main Location"
        self.Path = getattr(sys, '_MEIPASS', os.path.abspath(os.path.dirname(__file__)))
        self.Serial = self.Path + "\\Archives\\Serials.xlsx"  # Arquivo com os Nomes de Seriais.
        self.Screen() #Calls the TKInter Function that Generates the First Screen
        self.Frames() #Function that Generates Frames within Screen 01
        self.Widgets() #Screen that Generates Widgets
        self.Screen_Login() #Texts and Buttons
        self.Screen.mainloop() #Keep Screen 01 in the Loop

    #Definition of Screen 01 Attributes
    def Screen(self):
        self.Screen = Tk()
        self.Screen.title("Track RPA")  # Create Text in the Top Bar of the Window
        self.Screen.configure(background="#f8f8f8")  # Color in Background
        self.Screen.resizable(False, False)  # Screen Responsiveness. If the First Argument is False, it does not allow the Width to be changed. If the Second Argument is False, it does not allow the Height to be changed.

        # Determine the dimensions of the main window
        window_width = 700
        window_height = 500

        # Get screen dimensions
        screen_width = self.Screen.winfo_screenwidth()
        screen_height = self.Screen.winfo_screenheight()

        # Calculate window position in the center of the screen
        pos_x = (screen_width - window_width) // 2
        pos_y = (screen_height - window_height) // 2

        # Set the window geometry to open in the center of the screen
        self.Screen.geometry(f"{window_width}x{window_height}+{pos_x}+{pos_y}")

    #Frame Definition
    def Frames(self):
        #Frames in Screen 02
        self.FR2 = Frame(self.Screen, bd=0, highlightbackground="#000000", background='#072144', highlightthickness=2)
        self.FR2.place(relx=0, rely=0.95, relwidth=1.0, relheight=0.1)

    #Widgets Definition in Screen 01
    def Widgets(self):
        #Logo
        image_path = os.path.join(self.Path + "\\Archives\\images\\", "Logo_TrackTraceRX.png")
        self.Img = PhotoImage(file=image_path)
        self.Log = Label(self.Screen, image=self.Img, bg="#f8f8f8").place(relx=0.05, rely=0.09, relwidth=0.9, relheight=0.2)

        # Baseboard
        self.lb_base = Label(self.FR2, background='#072144', text="Powered By TrackTraceRX", fg='white', font=('Sans-serif', 8, 'bold')).place(relx=0.38, rely=0.01)

App()






